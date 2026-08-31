#!/usr/bin/env python3
"""Create an empty TSP Register.

Usage:
    python create_tsp.py "<output path>.xlsx" [--vocabulary en|pt]

Builds the five sheets the register needs — Tools Register, Control Activities,
Activity Log, Change Log, Lookups — with headers, styling and the controlled
vocabularies in place, and no data rows.

The existing personal register was migrated out of an Access database rather than
created by a script, so until now there was no way to stand up a fresh one. That
matters for running a second register somewhere else: the alternative was copying
a populated register and deleting every row, which risks carrying data across.

--vocabulary controls the Lookups values only. 'pt' reproduces the personal
register's historical Portuguese terms; 'en' is the English equivalent, for a
register starting from scratch. Relevancy and Frequency are English either way.
"""

import argparse
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")

SHEETS = [
    ("Tools Register", [
        ("ID", 6), ("Tool/System Name", 30), ("Description", 40), ("Type", 18),
        ("Status", 15), ("Relevancy", 18), ("Primary AF", 12), ("Other AFs", 14),
        ("Doc Aux", 10), ("Links", 30), ("Notes", 30), ("Last Reviewed", 14),
        ("Skill", 46),
    ]),
    ("Control Activities", [
        ("ID", 6), ("Activity Name", 45), ("Frequency", 14), ("Duration (min)", 14),
        ("Importance", 14), ("Commitment", 14), ("Linked Tool", 28),
        ("Description", 40), ("Last Done", 14), ("Next Due", 14),
    ]),
    ("Activity Log", [
        ("ID", 6), ("Activity", 45), ("Done On", 14), ("Planned For", 14),
        ("Notes", 40), ("Review On", 14), ("Times Postponed", 16),
    ]),
    ("Change Log", [
        ("ID", 6), ("Tool", 28), ("Changed On", 14), ("Description", 90),
    ]),
]

# Frequency carries a Days value; Next Due = Last Done + Days.
FREQUENCY = [("Daily", 1), ("Weekly", 7), ("Monthly", 31), ("6 weeks", 42),
             ("2 months", 60), ("Quarterly", 90), ("4 months", 120),
             ("Semi-annual", 180), ("Annual", 365), ("2-2 Years", 730),
             ("5-5 Years", 1825)]

RELEVANCY = ["Critical", "Often", "Sometimes", "Specific - relevant",
             "Specific - questionable", "Rarely", "Not in the last years"]

VOCABULARY = {
    "pt": {
        "Status": ["Planeado", "Em execução", "Implementado", "Absoleto"],
        "Type": ["Manual", "Instrução de trabalho", "Ferramenta"],
        "Importance": ["Crítica", "Importante", "Opcional",
                       "Não necessário", "Absoleto"],
    },
    "en": {
        "Status": ["Planned", "In Progress", "Implemented", "Obsolete"],
        "Type": ["Manual", "Work Instruction", "Tool"],
        "Importance": ["Critical", "Important", "Optional", "Not needed", "Obsolete"],
    },
}


def write_headers(ws, columns):
    for index, (title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s1" % get_column_letter(len(columns))


def write_lookups(ws, vocabulary):
    """Stacked blocks: a title row, an 'ID'/<column> header, values, blank row."""
    blocks = [
        ("Status Values", "Status", [(v,) for v in vocabulary["Status"]]),
        ("Type Values", "Type", [(v,) for v in vocabulary["Type"]]),
        ("Relevancy Values", "Relevancy", [(v,) for v in RELEVANCY]),
        ("Importance Values", "Importance", [(v,) for v in vocabulary["Importance"]]),
        ("Frequency Values", "Frequency", [(n, d) for n, d in FREQUENCY]),
    ]
    row = 1
    for title, column, values in blocks:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="ID")
        ws.cell(row=row, column=2, value=column)
        if column == "Frequency":
            ws.cell(row=row, column=3, value="Days")
        for cell in ws[row][:3]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row += 1
        for index, value in enumerate(values, start=1):
            ws.cell(row=row, column=1, value=index)
            ws.cell(row=row, column=2, value=value[0])
            if len(value) > 1:
                ws.cell(row=row, column=3, value=value[1])
            row += 1
        row += 1  # blank separator — the parser relies on it
    for letter, width in (("A", 6), ("B", 28), ("C", 8)):
        ws.column_dimensions[letter].width = width


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("output", help="path of the .xlsx to create")
    parser.add_argument("--vocabulary", choices=sorted(VOCABULARY), default="en",
                        help="controlled vocabulary for Status/Type/Importance (default: en)")
    parser.add_argument("--force", action="store_true",
                        help="overwrite the output file if it already exists")
    args = parser.parse_args()

    if os.path.exists(args.output) and not args.force:
        raise SystemExit("Refusing to overwrite an existing register: %s\n"
                         "Pass --force only if you are certain." % args.output)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, columns in SHEETS:
        write_headers(wb.create_sheet(name), columns)
    write_lookups(wb.create_sheet("Lookups"), VOCABULARY[args.vocabulary])
    wb.save(args.output)

    print("Created %s" % args.output)
    print("  sheets     %s" % ", ".join([n for n, _ in SHEETS] + ["Lookups"]))
    print("  vocabulary %s" % args.vocabulary)
    print("  data rows  0 - assign the first tool ID 1 and grow from there")
    return 0


if __name__ == "__main__":
    sys.exit(main())
