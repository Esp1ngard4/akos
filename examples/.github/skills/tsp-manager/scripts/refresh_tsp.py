#!/usr/bin/env python3
"""Regenerate the TSP Dashboard from the TSP Register.

Usage:
    python refresh_tsp.py "<path to TSP Register.xlsx>" [--out "<dashboard.html>"]

Reads the 'Tools Register', 'Control Activities' and 'Change Log' sheets, injects
them into templates/dashboard.html and writes the result next to the register
(as 'TSP Dashboard.html') unless --out says otherwise.

The dashboard is a static snapshot: rerun this after any edit to the register.
"""

import argparse
import datetime as dt
import json
import os
import sys

import openpyxl

SHEETS = {
    "TOOLS": "Tools Register",
    "ACTIVITIES": "Control Activities",
    "CHANGELOG": "Change Log",
}

TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir,
                        "templates", "dashboard.html")


def cell(value):
    """Normalise a cell to something JSON-serialisable and dashboard-friendly."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value.strip()
    return value


def read_sheet(wb, name):
    """Return a sheet as a list of dicts keyed by the row-1 headers."""
    if name not in wb.sheetnames:
        raise SystemExit("Register is missing the '%s' sheet." % name)
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    headers = [cell(h) for h in next(rows)]
    records = []
    for row in rows:
        # A blank ID means a spacer/trailing row, not a record.
        if row[0] is None or cell(row[0]) == "":
            continue
        record = {}
        for header, value in zip(headers, row):
            if header:
                record[header] = cell(value)
        records.append(record)
    return records


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("register", help="path to TSP Register.xlsx")
    parser.add_argument("--out", help="output html path (default: alongside the register)")
    args = parser.parse_args()

    if not os.path.isfile(args.register):
        raise SystemExit("Register not found: %s" % args.register)

    wb = openpyxl.load_workbook(args.register, data_only=True)
    data = {key: read_sheet(wb, sheet) for key, sheet in SHEETS.items()}

    if not os.path.isfile(TEMPLATE):
        raise SystemExit(
            "Dashboard template not found:\n  %s\n\n"
            "refresh_tsp.py renders the register into this template. It ships with\n"
            "the skill at <skill>/templates/dashboard.html - if it is missing, the\n"
            "skill was copied or packaged without it."
            % os.path.abspath(TEMPLATE))

    with open(TEMPLATE, encoding="utf-8") as fh:
        html = fh.read()

    for key, records in data.items():
        placeholder = "{{%s}}" % key
        if placeholder not in html:
            raise SystemExit("Template is missing the %s placeholder." % placeholder)
        html = html.replace(placeholder,
                            json.dumps(records, ensure_ascii=False, default=str))
    html = html.replace("{{GENERATED}}", dt.date.today().strftime("%Y-%m-%d"))

    out = args.out or os.path.join(os.path.dirname(os.path.abspath(args.register)),
                                   "TSP Dashboard.html")
    with open(out, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(html)

    print("Wrote %s" % out)
    for key, records in data.items():
        print("  %-11s %d rows" % (key.lower(), len(records)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
