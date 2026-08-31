#!/usr/bin/env python3
"""Cross-check the TSP Register against what is actually on disk.

Usage:
    python audit_tsp.py "<path to TSP Register.xlsx>" [--fsp-root "<tools folder>"]
                                                      [--skills "<dir>" ...]

Reports, in order:
  1. FSP folders with no register row (including unnumbered folders)
  2. Non-obsolete register rows with no FSP folder
  3. Rows overdue for their annual review (DF.18 control 2)
  4. Overdue control activities (DF.18 control 1)
  5. Values outside the Lookups sheet
  6. Skill column vs. what is actually installed, reconciled both ways

This is a read-only report. It never writes to the register.
"""

import argparse
import datetime as dt
import os
import re
import sys

import openpyxl

OBSOLETE = {"absoleto", "obsoleto", "obsolete"}
FOLDER_RE = re.compile(r"^FSP\.(\d+)\s+(.*)$")
ANNUAL_DAYS = 365


def as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return dt.datetime.strptime(value.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def load_lookups(wb):
    """Parse the Lookups sheet into {'Status': {...}, 'Type': {...}, ...}.

    The sheet is a stack of blocks: a title row, an 'ID'/<name> header row, then
    values until a blank row. The header's second cell is the column it governs.
    """
    lookups = {}
    if "Lookups" not in wb.sheetnames:
        return lookups
    current = None
    for row in wb["Lookups"].iter_rows(values_only=True):
        first, second = row[0], row[1] if len(row) > 1 else None
        if first == "ID" and second:
            current = str(second).strip()
            lookups[current] = set()
        elif current and isinstance(first, int) and second:
            lookups[current].add(str(second).strip())
        elif first is None and second is None:
            current = None
    return lookups


def load_register(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    tools, activities = [], []
    ws = wb["Tools Register"]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        tools.append(dict(zip(headers, row)))
    ws = wb["Control Activities"]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        activities.append(dict(zip(headers, row)))
    return tools, activities, load_lookups(wb)


def split_skills(value):
    """The Skill column holds a comma-separated list of skill names."""
    if not value:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def find_skill_roots(repo_root, fsp_root):
    """Locate the '.claude/skills' directories that hold this system's skills.

    Skills are not all in one place: the core set sits at the repo root, but
    some tools ship their own scoped skills inside their own tool folder.
    Project folders elsewhere in the repo hold project-scoped skills, which
    belong to projects rather than TSPs and are deliberately not scanned.

    Both roots are derived from paths the caller already resolved, so no folder
    name is hardcoded here.
    """
    roots = []
    top = os.path.join(repo_root, ".claude", "skills")
    if os.path.isdir(top):
        roots.append(top)
    for dirpath, dirnames, _ in os.walk(fsp_root):
        dirnames[:] = [d for d in dirnames
                       if d not in (".git", "node_modules", "PreviousV", "Other Versions")]
        if os.path.basename(dirpath) == "skills" \
                and os.path.basename(os.path.dirname(dirpath)) == ".claude":
            roots.append(dirpath)
    return roots


def scan_folders(fsp_root):
    numbered, unnumbered = {}, []
    for name in sorted(os.listdir(fsp_root)):
        if not os.path.isdir(os.path.join(fsp_root, name)):
            continue
        match = FOLDER_RE.match(name)
        if match:
            numbered[int(match.group(1))] = name
        elif name.startswith("FSP"):
            unnumbered.append(name)
    return numbered, unnumbered


def section(title, lines):
    print("\n== %s ==" % title)
    if not lines:
        print("  (none)")
    for line in lines:
        print("  " + line)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("register")
    parser.add_argument("--fsp-root", help="default: grandparent of the register")
    parser.add_argument("--skills", action="append", metavar="DIR",
                        help="a .claude/skills directory to scan; repeatable. "
                             "Default: auto-discovered (see find_skill_roots)")
    args = parser.parse_args()

    register = os.path.abspath(args.register)
    fsp_root = args.fsp_root or os.path.dirname(os.path.dirname(register))
    today = dt.date.today()

    tools, activities, lookups = load_register(register)
    numbered, unnumbered = scan_folders(fsp_root)
    by_id = {}
    for tool in tools:
        try:
            by_id[int(tool["ID"])] = tool
        except (TypeError, ValueError):
            continue

    print("Register: %s" % register)
    print("FSP root: %s" % fsp_root)
    print("%d register rows, %d numbered FSP folders" % (len(by_id), len(numbered)))

    section("FSP folders with no register row", [
        "%s" % name for fsp_id, name in sorted(numbered.items()) if fsp_id not in by_id
    ] + ["%s  (unnumbered)" % name for name in unnumbered])

    section("Non-obsolete register rows with no FSP folder", [
        "%-4s %-40s %s" % (fsp_id, by_id[fsp_id].get("Tool/System Name"),
                           by_id[fsp_id].get("Status"))
        for fsp_id in sorted(by_id)
        if fsp_id not in numbered
        and str(by_id[fsp_id].get("Status", "")).strip().lower() not in OBSOLETE
    ])

    overdue_review = []
    for fsp_id in sorted(by_id):
        tool = by_id[fsp_id]
        if str(tool.get("Status", "")).strip().lower() in OBSOLETE:
            continue
        reviewed = as_date(tool.get("Last Reviewed"))
        if reviewed is None:
            overdue_review.append("%-4s %-40s never reviewed" % (fsp_id, tool.get("Tool/System Name")))
        elif (today - reviewed).days > ANNUAL_DAYS:
            overdue_review.append("%-4s %-40s last %s (%d days)" % (
                fsp_id, tool.get("Tool/System Name"), reviewed, (today - reviewed).days))
    section("Overdue annual reviews (%d)" % len(overdue_review), overdue_review)

    overdue_acts = []
    for act in activities:
        due = as_date(act.get("Next Due"))
        if due and due < today:
            overdue_acts.append((due, "%-4s %-45s due %s (%d days)" % (
                act.get("ID"), str(act.get("Activity Name"))[:45], due, (today - due).days)))
    overdue_acts.sort()
    section("Overdue control activities (%d)" % len(overdue_acts),
            [line for _, line in overdue_acts])

    # (sheet, records, label column, columns whose values the Lookups sheet governs)
    checks = [
        ("Tools Register", tools, "Tool/System Name", ["Status", "Type", "Relevancy"]),
        ("Control Activities", activities, "Activity Name", ["Frequency", "Importance"]),
    ]
    bad_values = []
    for sheet, records, label, columns in checks:
        for record in records:
            for column in columns:
                allowed = lookups.get(column)
                value = record.get(column)
                if not allowed or value is None or str(value).strip() == "":
                    continue
                if str(value).strip() not in allowed:
                    bad_values.append("%-18s %-4s %-32s %-11s = %r" % (
                        sheet, record.get("ID"), str(record.get(label))[:32], column, value))
    section("Values outside the Lookups sheet (%d)" % len(bad_values), bad_values)

    roots = args.skills or find_skill_roots(os.path.dirname(fsp_root), fsp_root)
    installed = {}
    for root in roots:
        for name in sorted(os.listdir(root)):
            if os.path.isfile(os.path.join(root, name, "SKILL.md")):
                installed.setdefault(name, root)

    claimed = {}
    for fsp_id in sorted(by_id):
        for name in split_skills(by_id[fsp_id].get("Skill")):
            claimed.setdefault(name, []).append(fsp_id)

    print("\nSkill roots scanned:")
    for root in roots:
        print("  %s" % root)

    section("Skills named in the register but not installed", [
        "%-34s claimed by FSP.%s" % (name, ", FSP.".join(str(i) for i in ids))
        for name, ids in sorted(claimed.items()) if name not in installed])

    section("Skills installed but claimed by no register row", [
        "%-34s %s" % (name, installed[name])
        for name in sorted(installed) if name not in claimed])

    automated = [fsp_id for fsp_id in sorted(by_id) if by_id[fsp_id].get("Skill")]
    section("Tool-to-skill map (%d tools)" % len(automated), [
        "FSP.%-4s %-34s %s" % (fsp_id, str(by_id[fsp_id].get("Tool/System Name"))[:34],
                               by_id[fsp_id].get("Skill"))
        for fsp_id in automated])

    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
