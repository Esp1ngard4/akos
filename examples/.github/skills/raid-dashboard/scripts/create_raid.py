#!/usr/bin/env python3
"""
Create a fresh RAID .xlsx register for a new project.
Usage: python create_raid.py <output_path> [project_name]
"""
import sys, os
from datetime import datetime

def create_raid_xlsx(output_path, project_name="New Project"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket Tracker"

    # Title rows
    ws.merge_cells("B2:G2")
    ws["B2"] = f"RAID Register — {project_name}"
    ws["B2"].font = Font(size=16, bold=True, color="1F2937")

    ws.merge_cells("B3:G3")
    ws["B3"] = f"Created: {datetime.now().strftime('%Y-%m-%d')}"
    ws["B3"].font = Font(size=10, color="6B7280")

    # Column definitions (v2.3 schema — risk-analysis block inserted after Feasibility,
    # Next Review On after Review On, ETC Renegotiated after ETC — see WP2.1 Change Spec)
    headers = {
        2: ("RAID.ID", 10),
        3: ("Detail", 30),
        4: ("Type", 12),
        5: ("DRI", 14),
        6: ("Priority %", 12),
        7: ("Urgency (1-5)", 12),
        8: ("Consequences (1-5)", 14),
        9: ("Feasibility", 12),
        10: ("Probability of Occurrence (1-5)", 16),
        11: ("Severity (1-5)", 12),
        12: ("Response Strategy", 16),
        13: ("Mitigation Target %", 16),
        14: ("Target Residual Risk", 16),
        15: ("Residual Risk Score", 16),
        16: ("MoSCoW", 14),
        17: ("Status", 12),
        18: ("Last Review", 14),
        19: ("Review On", 14),
        20: ("Next Review On", 14),
        21: ("Description", 40),
        22: ("Action Plan", 35),
        23: ("Acceptance Criteria", 30),
        24: ("Action Log", 35),
        25: ("Category", 14),
        26: ("Tracked Externally", 16),
        27: ("Opened On", 14),
        28: ("Requested By", 14),
        29: ("Involve", 18),
        30: ("Has AuxMat", 14),
        31: ("Estimated Effort", 14),
        32: ("ETC", 12),
        33: ("ETC Renegotiated", 14),
        34: ("Closed On", 14),
        35: ("Closed By", 14),
    }

    header_row = 6
    header_font = Font(size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col, (name, width) in headers.items():
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data validation lists
    from openpyxl.worksheet.datavalidation import DataValidation

    type_dv = DataValidation(type="list", formula1='"Risk,Action,Issue,Decision,Idea"', allow_blank=True)
    type_dv.error = "Please select a valid type"
    ws.add_data_validation(type_dv)
    type_dv.add(f"D{header_row+1}:D{header_row+100}")

    strategy_dv = DataValidation(type="list", formula1='"Avoid,Transfer,Mitigate,Accept,Exploit,Share,Enhance"', allow_blank=True)
    strategy_dv.error = "Please select a valid response strategy"
    ws.add_data_validation(strategy_dv)
    strategy_dv.add(f"L{header_row+1}:L{header_row+100}")

    moscow_dv = DataValidation(type="list", formula1='"1.Must,2.Should,3.Could,4.Wont"', allow_blank=True)
    ws.add_data_validation(moscow_dv)
    moscow_dv.add(f"P{header_row+1}:P{header_row+100}")

    status_dv = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Closed,On Hold"', allow_blank=True)
    ws.add_data_validation(status_dv)
    status_dv.add(f"Q{header_row+1}:Q{header_row+100}")

    tracked_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    tracked_dv.error = "Please enter Y or N"
    ws.add_data_validation(tracked_dv)
    tracked_dv.add(f"Z{header_row+1}:Z{header_row+100}")

    auxmat_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    auxmat_dv.error = "Please enter Y or N"
    ws.add_data_validation(auxmat_dv)
    auxmat_dv.add(f"AD{header_row+1}:AD{header_row+100}")

    # Formulas for first 100 rows: Priority % (F) and Target Residual Risk (N)
    for r in range(header_row + 1, header_row + 101):
        ws.cell(row=r, column=6).value = f'=IF(OR(G{r}="",H{r}=""),(G{r}*1.5+H{r})/12.5*100,ROUND((G{r}*1.5+H{r})/12.5*100,1))'
        ws.cell(row=r, column=6).number_format = '0.0"%"'
        ws.cell(row=r, column=13).number_format = '0"%"'
        ws.cell(row=r, column=14).value = f'=IF(M{r}="","",ROUND(J{r}*K{r}*(1-M{r}/100),1))'
        ws.cell(row=r, column=14).number_format = '0.0'

    # Light alternating row fills for readability
    light_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    for r in range(header_row + 1, header_row + 101):
        for col in headers:
            cell = ws.cell(row=r, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = light_fill

    # Freeze panes at header
    ws.freeze_panes = f"B{header_row + 1}"

    wb.save(output_path)

    # Create AuxMat folder alongside the register
    auxmat_path = os.path.join(os.path.dirname(output_path), "AuxMat")
    os.makedirs(auxmat_path, exist_ok=True)

    print(f"RAID register created: {output_path}")
    print(f"AuxMat folder created: {auxmat_path}")
    print(f"Project: {project_name}")
    print(f"Schema: {len(headers)} columns (risk-analysis block J-O), priority formula in col F")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python create_raid.py <output_path> [project_name]")
        sys.exit(1)
    output_path = sys.argv[1]
    project_name = sys.argv[2] if len(sys.argv) > 2 else "New Project"
    create_raid_xlsx(output_path, project_name)
