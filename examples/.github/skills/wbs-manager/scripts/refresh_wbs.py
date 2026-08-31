#!/usr/bin/env python3
"""
WBS Dashboard Generator (AI-First)
Reads a WBS xlsx, exports a self-contained HTML dashboard.

Usage:
    python refresh_wbs.py "<xlsx-path>" "<output-html-path>" "<project-name>"
"""

import sys
import json
import openpyxl
from pathlib import Path
from datetime import datetime


def discover_schema(ws):
    """Build column map from header row 1."""
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and str(v).strip():
            col_map[str(v).strip()] = c
    return col_map


def extract_items(ws, col_map):
    """Extract all data rows from a WBS sheet."""
    items = []
    for r in range(2, ws.max_row + 1):
        item = {}
        has_data = False
        for name, ci in col_map.items():
            val = ws.cell(row=r, column=ci).value
            if val is not None:
                item[name] = val if isinstance(val, (int, float)) else str(val)
                if name in ('Code', 'Title') and str(val).strip():
                    has_data = True
        if has_data:
            items.append(item)
    return items


def compute_stats(items):
    """Compute summary statistics."""
    stats = {
        'total': len(items),
        'done': 0, 'implementing': 0, 'not_started': 0,
        'backlog': 0, 'funnel': 0, 'no_status': 0,
        'total_effort': 0,
        'sprints': {},
        'statuses': {},
        'priorities': {},
        'types': {},
    }
    for it in items:
        s = str(it.get('Status', 'No Status') or 'No Status')
        stats['statuses'][s] = stats['statuses'].get(s, 0) + 1
        if s == 'Done': stats['done'] += 1
        elif s == 'Implementing': stats['implementing'] += 1
        elif s == 'Not Started': stats['not_started'] += 1
        elif s == 'Portfolio Backlog': stats['backlog'] += 1
        elif s == 'Funnel': stats['funnel'] += 1
        else: stats['no_status'] += 1

        p = str(it.get('Priority', 'None') or 'None')
        stats['priorities'][p] = stats['priorities'].get(p, 0) + 1

        t = str(it.get('Type', '') or '')
        if t:
            stats['types'][t] = stats['types'].get(t, 0) + 1

        stats['total_effort'] += it.get('Estimated Effort (h)', 0) or 0

        sp = it.get('Sprint Planned')
        if sp and str(sp).startswith('S'):
            sp = str(sp)
            if sp not in stats['sprints']:
                stats['sprints'][sp] = {'items': [], 'effort': 0, 'done': 0}
            stats['sprints'][sp]['items'].append(it)
            stats['sprints'][sp]['effort'] += it.get('Estimated Effort (h)', 0) or 0
            if s == 'Done':
                stats['sprints'][sp]['done'] += 1

    return stats


def generate_html(project_name, items, stats):
    """Generate self-contained HTML dashboard with filters."""

    compact = []
    for it in items:
        compact.append({
            'c': str(it.get('Code', '')),
            't': str(it.get('Title', '')),
            's': str(it.get('Status', 'No Status') or 'No Status'),
            'p': str(it.get('Priority', '') or ''),
            'e': it.get('Estimated Effort (h)', 0) or 0,
            'sp': str(it.get('Sprint Planned', '') or ''),
            'se': str(it.get('Sprint Ended', '') or ''),
            'ty': str(it.get('Type', '') or ''),
            'cat': str(it.get('Category', '') or ''),
            'o': str(it.get('Owner', '') or ''),
        })

    items_json = json.dumps(compact)
    stats_json = json.dumps({
        'total': stats['total'], 'done': stats['done'],
        'implementing': stats['implementing'], 'not_started': stats['not_started'],
        'backlog': stats['backlog'], 'funnel': stats['funnel'], 'no_status': stats['no_status'],
        'total_effort': stats['total_effort'],
        'statuses': stats['statuses'], 'priorities': stats['priorities'],
        'types': stats['types'],
    })

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{project_name} — WBS Dashboard</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,-apple-system,sans-serif;color:#1a1a2e;background:#fff;padding:24px;max-width:1200px;margin:0 auto}}
.header{{margin-bottom:24px;border-bottom:3px solid #4472C4;padding-bottom:12px}}
.header h1{{font-size:22px;color:#4472C4}}
.header .sub{{font-size:13px;color:#6b7280;margin-top:4px}}
.kpi-strip{{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap}}
.kpi{{flex:1;min-width:110px;padding:14px 16px;border-radius:8px;border:1px solid #e5e7eb}}
.kpi-val{{font-size:28px;font-weight:700;line-height:1.2}}
.kpi-label{{font-size:12px;color:#6b7280;margin-top:2px}}
.tabs{{display:flex;gap:0;border-bottom:2px solid #e5e7eb;margin-bottom:16px}}
.tab{{padding:8px 18px;cursor:pointer;font-size:13px;font-weight:600;color:#6b7280;border-bottom:2px solid transparent;margin-bottom:-2px;transition:.15s;user-select:none}}
.tab:hover{{color:#1a1a2e}}
.tab.active{{color:#4472C4;border-bottom-color:#4472C4}}
.panel{{display:none}}.panel.active{{display:block}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th{{background:#4472C4;color:#fff;padding:8px 6px;text-align:left;position:sticky;top:0;font-weight:600}}
td{{padding:6px;border-bottom:1px solid #e5e7eb;vertical-align:top}}
tr:hover td{{background:#f0f4ff}}
.badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}}
.b-done{{background:#dcfce7;color:#166534}}.b-impl{{background:#dbeafe;color:#1e40af}}
.b-ns{{background:#f3f4f6;color:#374151}}.b-bl{{background:#fef3c7;color:#92400e}}
.b-fn{{background:#fce7f3;color:#9d174d}}.b-no{{background:#f9fafb;color:#9ca3af}}
.sprint-section{{margin-bottom:24px}}
.sprint-header{{display:flex;align-items:center;gap:10px;padding:10px 0;border-bottom:2px solid #4472C4;margin-bottom:8px}}
.sprint-name{{font-size:16px;font-weight:700}}
.sprint-meta{{font-size:12px;color:#6b7280}}
.chart-row{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:20px}}
.chart-box{{flex:1;min-width:220px;padding:16px;border:1px solid #e5e7eb;border-radius:8px}}
.chart-title{{font-size:13px;font-weight:600;margin-bottom:12px}}
.bar-row{{display:flex;align-items:center;gap:8px;margin-bottom:6px}}
.bar-label{{width:90px;font-size:11px;text-align:right;flex-shrink:0}}
.bar-track{{flex:1;height:22px;background:#e5e7eb;border-radius:4px;overflow:hidden}}
.bar-fill{{height:100%;border-radius:4px;display:flex;align-items:center;padding:0 6px;font-size:10px;color:#fff;font-weight:600;min-width:fit-content}}
.gantt-row{{display:flex;align-items:center;margin-bottom:4px}}
.gantt-label{{width:220px;font-size:11px;flex-shrink:0;padding-right:8px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.gantt-track{{flex:1;height:24px;position:relative;background:#f3f4f6;border-radius:3px}}
.gantt-bar{{position:absolute;height:20px;top:2px;border-radius:3px;display:flex;align-items:center;padding:0 6px;font-size:9px;color:#fff;font-weight:600;white-space:nowrap;overflow:hidden}}
.legend{{margin-bottom:12px;font-size:12px;color:#6b7280}}
.legend span{{display:inline-block;width:12px;height:12px;border-radius:2px;vertical-align:middle;margin-right:4px}}
.filters{{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}}
.filters select{{padding:5px 10px;border:1px solid #e5e7eb;border-radius:4px;font-size:12px;background:#fff;color:#1a1a2e;cursor:pointer}}
.filters select:focus{{outline:none;border-color:#4472C4}}
.timestamp{{margin-top:24px;padding-top:12px;border-top:1px solid #e5e7eb;font-size:11px;color:#6b7280;text-align:right}}
</style>
</head>
<body>
<div class="header">
  <h1>{project_name} — WBS Dashboard</h1>
  <div class="sub">AI-First Project Management · Generated {timestamp}</div>
</div>
<div id="app"></div>
<script>
const DATA={items_json};
const S={stats_json};

let fStatus='All',fPriority='All',fType='All';

const sprintGroups={{}};
function rebuildSprints(){{
  Object.keys(sprintGroups).forEach(k=>delete sprintGroups[k]);
  filtered().filter(d=>d.sp&&d.sp.startsWith('S')).forEach(d=>{{
    if(!sprintGroups[d.sp])sprintGroups[d.sp]={{items:[],effort:0,done:0}};
    sprintGroups[d.sp].items.push(d);
    sprintGroups[d.sp].effort+=d.e||0;
    if(d.s==='Done')sprintGroups[d.sp].done++;
  }});
}}

function filtered(){{
  return DATA.filter(d=>{{
    if(fStatus!=='All'&&d.s!==fStatus)return false;
    if(fPriority!=='All'&&d.p!==fPriority)return false;
    if(fType!=='All'&&d.ty!==fType)return false;
    return true;
  }});
}}

function unique(key){{
  const vals=new Set();DATA.forEach(d=>{{if(d[key])vals.add(d[key]);}});
  return [...vals].sort();
}}

function badge(s){{
  const m={{'Done':'b-done','Implementing':'b-impl','Not Started':'b-ns','Portfolio Backlog':'b-bl','Funnel':'b-fn'}};
  return `<span class="badge ${{m[s]||'b-no'}}">${{s}}</span>`;
}}

function bars(data,mx){{
  return data.map(d=>{{
    const w=mx>0?Math.round(d.v/mx*100):0;
    return `<div class="bar-row"><div class="bar-label">${{d.l}}</div><div class="bar-track"><div class="bar-fill" style="width:${{Math.max(w,8)}}%;background:${{d.c}}">${{d.v}}</div></div></div>`;
  }}).join('');
}}

function renderFilters(){{
  const statuses=unique('s'),priorities=unique('p'),types=unique('ty');
  return `<div class="filters">
    <select onchange="fStatus=this.value;render()"><option value="All">All Statuses</option>${{statuses.map(s=>`<option value="${{s}}" ${{fStatus===s?'selected':''}}>${{s}}</option>`).join('')}}</select>
    <select onchange="fPriority=this.value;render()"><option value="All">All Priorities</option>${{priorities.map(p=>`<option value="${{p}}" ${{fPriority===p?'selected':''}}>${{p}}</option>`).join('')}}</select>
    <select onchange="fType=this.value;render()"><option value="All">All Types</option>${{types.map(t=>`<option value="${{t}}" ${{fType===t?'selected':''}}>${{t}}</option>`).join('')}}</select>
  </div>`;
}}

function renderKPIs(){{
  return `<div class="kpi-strip">
    <div class="kpi"><div class="kpi-val">${{S.total}}</div><div class="kpi-label">Total Items</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#3b82f6">${{S.implementing}}</div><div class="kpi-label">Implementing</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#22c55e">${{S.done}}</div><div class="kpi-label">Done</div></div>
    <div class="kpi"><div class="kpi-val" style="color:#f59e0b">${{S.not_started}}</div><div class="kpi-label">Not Started</div></div>
    <div class="kpi"><div class="kpi-val">${{S.total_effort}}h</div><div class="kpi-label">Total Effort</div></div>
    <div class="kpi"><div class="kpi-val">${{Object.keys(sprintGroups).length}}</div><div class="kpi-label">Planned Sprints</div></div>
  </div>`;
}}

function renderBoard(){{
  let h='';
  Object.keys(sprintGroups).sort().forEach(sp=>{{
    const g=sprintGroups[sp];
    const pct=g.items.length>0?Math.round(g.done/g.items.length*100):0;
    h+=`<div class="sprint-section"><div class="sprint-header"><span class="sprint-name">${{sp}}</span><span class="sprint-meta">${{g.items.length}} items · ${{g.effort}}h · ${{pct}}% done</span></div>
    <table><tr><th>Code</th><th>Title</th><th>Status</th><th>Priority</th><th>Effort</th><th>Type</th><th>Owner</th></tr>`;
    g.items.forEach(d=>{{h+=`<tr><td>${{d.c}}</td><td>${{d.t}}</td><td>${{badge(d.s)}}</td><td>${{d.p||'-'}}</td><td>${{d.e||'-'}}h</td><td>${{d.ty||'-'}}</td><td>${{d.o||'-'}}</td></tr>`;}});
    h+=`</table></div>`;
  }});
  const f=filtered();
  const unplanned=f.filter(d=>(!d.sp||!d.sp.startsWith('S'))&&d.p==='Must'&&d.s!=='Done');
  if(unplanned.length){{
    h+=`<div class="sprint-section"><div class="sprint-header"><span class="sprint-name">Unplanned Must-Haves</span><span class="sprint-meta">${{unplanned.length}} items</span></div>
    <table><tr><th>Code</th><th>Title</th><th>Status</th><th>Effort</th><th>Type</th><th>Owner</th></tr>`;
    unplanned.forEach(d=>{{h+=`<tr><td>${{d.c}}</td><td>${{d.t}}</td><td>${{badge(d.s)}}</td><td>${{d.e||'-'}}h</td><td>${{d.ty||'-'}}</td><td>${{d.o||'-'}}</td></tr>`;}});
    h+=`</table></div>`;
  }}
  if(!h)h='<p style="color:#6b7280;padding:20px 0">No items match the current filters.</p>';
  return h;
}}

function renderAnalytics(){{
  const sc={{'Done':'#22c55e','Implementing':'#3b82f6','Not Started':'#f59e0b','Portfolio Backlog':'#8b5cf6','Funnel':'#ec4899'}};
  const sd=Object.entries(S.statuses).map(([k,v])=>({{l:k,v,c:sc[k]||'#9ca3af'}}));
  const pc={{'Must':'#ef4444','Should':'#f59e0b','Could':'#22c55e'}};
  const pd=Object.entries(S.priorities).map(([k,v])=>({{l:k,v,c:pc[k]||'#9ca3af'}}));
  const tc=['#4472C4','#22c55e','#f59e0b','#8b5cf6','#ec4899','#ef4444','#06b6d4','#84cc16'];
  const td=Object.entries(S.types).map(([k,v],i)=>({{l:k,v,c:tc[i%tc.length]}}));
  const spd=Object.keys(sprintGroups).sort().map((sp,i)=>({{l:sp,v:sprintGroups[sp].effort,c:['#4472C4','#22c55e','#f59e0b','#8b5cf6','#ec4899'][i%5]}}));
  return `<div class="chart-row"><div class="chart-box"><div class="chart-title">Status Distribution</div>${{bars(sd,Math.max(...sd.map(d=>d.v),1))}}</div>
  <div class="chart-box"><div class="chart-title">Priority Breakdown</div>${{bars(pd,Math.max(...pd.map(d=>d.v),1))}}</div></div>
  <div class="chart-row"><div class="chart-box"><div class="chart-title">Type Distribution</div>${{td.length?bars(td,Math.max(...td.map(d=>d.v),1)):'<p style="color:#6b7280">No types assigned</p>'}}</div>
  <div class="chart-box"><div class="chart-title">Sprint Effort Allocation</div>${{spd.length?bars(spd,Math.max(...spd.map(d=>d.v),1)):'<p style="color:#6b7280">No sprints planned</p>'}}</div></div>`;
}}

function renderGantt(){{
  const sprints=Object.keys(sprintGroups).sort();
  if(!sprints.length)return '<p style="color:#6b7280">No sprint-planned items to show.</p>';
  const colors=['#4472C4','#22c55e','#f59e0b','#8b5cf6','#ec4899'];
  let h=`<div class="legend"><span style="background:#86efac"></span> Done <span style="background:#4472C4;margin-left:12px"></span> Planned</div>`;
  h+=`<div class="gantt-row" style="margin-bottom:8px"><div class="gantt-label" style="font-weight:700">Work Item</div><div class="gantt-track" style="display:flex;background:transparent">`;
  sprints.forEach(sp=>{{h+=`<div style="flex:1;text-align:center;font-size:11px;font-weight:600;padding:4px;border-bottom:2px solid #4472C4">${{sp}}</div>`;}});
  h+=`</div></div>`;
  const allSI=filtered().filter(d=>sprints.includes(d.sp));
  allSI.forEach(d=>{{
    const idx=sprints.indexOf(d.sp);
    const left=(idx/sprints.length*100),width=(1/sprints.length*100);
    const color=d.s==='Done'?'#86efac':colors[idx%colors.length];
    const tc=d.s==='Done'?'#166534':'#fff';
    h+=`<div class="gantt-row"><div class="gantt-label" title="${{d.t}}">${{d.c}}. ${{d.t}}</div><div class="gantt-track"><div class="gantt-bar" style="left:${{left}}%;width:${{width}}%;background:${{color}};color:${{tc}}">${{d.e}}h</div></div></div>`;
  }});
  return h;
}}

let activeTab='sprints';
function render(){{
  rebuildSprints();
  document.getElementById('app').innerHTML=`${{renderKPIs()}}
  <div class="tabs">
    <div class="tab ${{activeTab==='sprints'?'active':''}}" onclick="activeTab='sprints';render()">Sprint Board</div>
    <div class="tab ${{activeTab==='analytics'?'active':''}}" onclick="activeTab='analytics';render()">Analytics</div>
    <div class="tab ${{activeTab==='gantt'?'active':''}}" onclick="activeTab='gantt';render()">Gantt</div>
  </div>
  ${{renderFilters()}}
  <div class="panel active">${{activeTab==='sprints'?renderBoard():activeTab==='analytics'?renderAnalytics():renderGantt()}}</div>
  <div class="timestamp">Generated: {timestamp}</div>`;
}}
render();
</script>
</body>
</html>'''


def main():
    if len(sys.argv) < 4:
        print("Usage: refresh_wbs.py <xlsx-path> <output-html-path> <project-name>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    html_path = sys.argv[2]
    project_name = sys.argv[3]

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if name == 'WBS' or 'Implement' in name:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    col_map = discover_schema(ws)
    items = extract_items(ws, col_map)
    stats = compute_stats(items)
    html = generate_html(project_name, items, stats)

    Path(html_path).write_text(html, encoding='utf-8')
    print(f"Dashboard generated: {html_path}")
    print(f"  Items: {stats['total']}, Done: {stats['done']}, Sprints: {len(stats['sprints'])}")


if __name__ == '__main__':
    main()
