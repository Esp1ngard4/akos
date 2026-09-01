#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Check an Artifact Register against itself, against the TSP register, and against the disk.

    python audit_artifact_register.py "Artifact Register.xlsx"
    python audit_artifact_register.py "7. Artifact Register, Atlas.xlsx" \
        --root "<the folder it describes>" \
        --tsp-register "TSP Register.xlsx"

The filesystem check is the one no other register in this system can do. Because
the ID is written onto the artifact, the register makes a claim about the disk
that can be verified: a file with no ID prefix is unregistered, and a row whose
ID appears nowhere on disk has lost its artifact.

The same findings appear on the dashboard's Findings tab - both come from
`checks.py`, so they cannot disagree.

Exits non-zero if any error-level finding is present.
"""
import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checks                                                   # noqa: E402
import schema as S                                              # noqa: E402



def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("register")
    p.add_argument("--root", help="folder this register describes, for the disk check")
    p.add_argument("--tsp-register", help="the TSP register, for the Managed By check")
    p.add_argument("--summary", action="store_true",
                   help="class names and counts only, without the individual rows")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.register, data_only=True)
    ws = wb[S.SHEET] if S.SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    hrow, cols = S.find_header(ws)
    rows = S.read_rows(ws, hrow, cols)

    print("Artifact Register audit - %s" % os.path.basename(args.register))
    print("  %d rows, header on row %d" % (len(rows), hrow))

    errors, warnings, stats = checks.run(
        rows, root=args.root, tools=checks.load_tools(args.tsp_register),
        width=S.id_width(wb))

    if stats.get("delegations_resolved") is not None:
        print("  delegations: %d checked against the TSP register, %d resolved"
              % (stats["delegated"], stats["delegations_resolved"]))
    if stats.get("disk"):
        d = stats["disk"]
        print("  disk: %d prefixed entries, %d without an ID prefix, %d boundaries "
              "not descended into" % (d["prefixed"], d["bare"], d["boundaries"]))

    print()
    for msg in errors:
        print("  ERROR    %s" % msg)
    total = 0
    for kind in sorted(warnings, key=lambda k: -len(warnings[k])):
        items = warnings[kind]
        total += len(items)
        print("  warning  %s (%d)" % (kind, len(items)))
        if not args.summary:
            for msg in items:
                print("             %s" % msg)

    print("\n%d error(s), %d warning(s) in %d class(es)."
          % (len(errors), total, len(warnings)))
    return 1 if errors else 0


if __name__ == "__main__":
    sys.exit(main())
