# -*- coding: utf-8 -*-
"""The Artifact Register's health checks.

Defined once and called by both `audit_artifact_register.py`, which prints them,
and `refresh_artifact_register.py`, which renders them into the dashboard's
Findings tab. A check that lived in only one of those would be a check the other
silently lacked.
"""
import datetime as dt

import schema as S

STALE_YEARS = 2


def _stop_ids(rows, claimed):
    """IDs the disk scan must not descend into - the delegation rule, applied to
    the filesystem. See TD.5."""
    stop = set()
    for rec in rows:
        rid = S.clean(rec["ID"])
        kind = S.clean(rec.get("Type"))
        if S.clean(rec.get("Managed By")) or kind == "Tool" or rid not in claimed:
            stop.add(rid)
    return stop


def run(rows, root=None, tools=None, stale_years=STALE_YEARS):
    """Check a register. Returns (errors, warnings, stats).

    `root`  - folder the register describes; enables the disk reconciliation.
    `tools` - {id: name} from TSP.3; enables the delegation check.
    """
    errors = []
    warnings = {}
    stats = {}

    def err(msg):
        errors.append(msg)

    def warn(kind, msg):
        warnings.setdefault(kind, []).append(msg)

    ids = {}
    for rec in rows:
        rid = S.clean(rec["ID"])
        if rid in ids:
            err("duplicate ID %s (rows %d and %d)" % (rid, ids[rid], rec["_row"]))
        else:
            ids[rid] = rec["_row"]

    claimed = set()
    for rec in rows:
        digital, physical = S.parent_of(rec)
        claimed.update(p for p in (digital, physical) if p and p != S.ROOT_PARENT)

    for rec in rows:
        rid = S.clean(rec["ID"])
        where = "row %d (ID %s)" % (rec["_row"], rid)
        typ = S.clean(rec.get("Type"))
        status = S.clean(rec.get("Status"))

        if typ and typ not in S.TYPES:
            err("%s: Type %r not in %s" % (where, typ, S.TYPES))
        if status and status not in S.STATUSES:
            err("%s: Status %r not in %s" % (where, status, S.STATUSES))
        if not status:
            warn("no Status", where)
        if not typ:
            warn("no Type", where)

        for label, parent in zip(("Parent Digital", "Parent Physical"), S.parent_of(rec)):
            if parent and parent != S.ROOT_PARENT and parent not in ids:
                err("%s: %s points at %r, which is not an ID in this register"
                    % (where, label, parent))
            if parent and parent == rid:
                err("%s: %s points at itself" % (where, label))

        reviewed = rec.get("Last Reviewed")
        if isinstance(reviewed, dt.datetime) and status == "Active":
            age = (dt.datetime.now() - reviewed).days / 365.25
            if age > stale_years:
                warn("review older than %d years" % stale_years,
                     "%s last reviewed %s (%.1f yrs)" % (where, reviewed.date(), age))
        elif status == "Active" and not reviewed:
            warn("Active but never reviewed", where)

    # --- delegation ------------------------------------------------------
    delegated = [(S.clean(r["ID"]), S.clean(r.get("Managed By")), r["_row"])
                 for r in rows if S.clean(r.get("Managed By"))]
    stats["delegated"] = len(delegated)
    if tools is None:
        if delegated:
            warn("delegations unchecked",
                 "%d found - supply the TSP register to verify them" % len(delegated))
        stats["delegations_resolved"] = None
    else:
        resolved = 0
        for rid, target, row in delegated:
            key = target.replace(S.TOOL_PREFIX, "").strip()
            if key not in tools:
                err("row %d (ID %s): Managed By %r is not a tool in the TSP register"
                    % (row, rid, target))
            else:
                resolved += 1
                if not target.startswith(S.TOOL_PREFIX):
                    warn("Managed By not written %sn" % S.TOOL_PREFIX,
                         "row %d (ID %s): %r" % (row, rid, target))
        stats["delegations_resolved"] = resolved

    # --- the disk --------------------------------------------------------
    if root:
        stop = _stop_ids(rows, claimed)
        prefixed, bare = S.scan_folder(root, stop)
        stats["disk"] = {"prefixed": len(prefixed), "bare": len(bare),
                         "boundaries": len(stop), "root": root}
        for name in bare:
            err("unregistered on disk: %r has no ID prefix" % name)
        for pid, names in sorted(prefixed.items(), key=lambda kv: int(kv[0])):
            if pid not in ids:
                err("orphan on disk: %s is prefixed %s, which is not in the register"
                    % (names[0], pid))
            if len(names) > 1:
                nested = all(n.startswith(names[0] + "/") for n in names[1:])
                if nested:
                    warn("ID carried by both a folder and its contents",
                         "ID %s: %s" % (pid, ", ".join(names)))
                else:
                    err("ID %s is used by %d unrelated entries on disk: %s"
                        % (pid, len(names), names))
        for rec in rows:
            rid = S.clean(rec["ID"])
            digital, _ = S.parent_of(rec)
            if (digital and rid not in prefixed
                    and S.clean(rec.get("Status")) != "Retired"):
                warn("in the register but not on disk",
                     "row %d (ID %s) %s" % (rec["_row"], rid, S.clean(rec.get("Name"))))
    else:
        stats["disk"] = None
        warn("disk not checked",
             "supply the folder this register describes to reconcile against it")

    return errors, warnings, stats


def load_tools(path):
    """{id: name} from the TSP.3 tools register, or None."""
    import os
    import openpyxl
    if not path or not os.path.isfile(path):
        return None
    ws = openpyxl.load_workbook(path, data_only=True)["Tools Register"]
    tools = {}
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(r, 1).value
        if tid is not None:
            tools[str(tid).strip()] = ws.cell(r, 2).value
    return tools
