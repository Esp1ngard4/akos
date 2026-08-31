#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate the Artifact Register dashboard.

    python refresh_artifact_register.py "Artifact Register.xlsx" "Artifact Dashboard.html"
    python refresh_artifact_register.py <register> <output.html> --scope "Atlas"

The dashboard is derived output: overwritten in full on every run, never edited
by hand. The HTML is built here rather than read from a template file, so there
is no shipped asset to go missing from a copy of the skill.
"""
import argparse
import datetime as dt
import io
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checks                                                   # noqa: E402
import schema as S                                              # noqa: E402

STALE_YEARS = 2

TEMPLATE = u"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Artifact Register - {{SCOPE}}</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/gridjs@5.0.2/dist/theme/mermaid.min.css" crossorigin="anonymous">
<script src="https://cdn.jsdelivr.net/npm/gridjs@5.0.2/dist/gridjs.umd.js" crossorigin="anonymous"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.0/dist/chart.umd.js" crossorigin="anonymous"></script>
<style>
:root { color-scheme: light; }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f8f9fa; color: #1a1a2e; padding: 16px; }
.header { display: flex; align-items: center; gap: 12px; margin-bottom: 20px; }
.header h1 { font-size: 22px; font-weight: 700; color: #2F5496; }
.header .subtitle { color: #666; font-size: 13px; }
.tabs { display: flex; gap: 4px; margin-bottom: 16px; border-bottom: 2px solid #e0e0e0; }
.tab { padding: 8px 16px; cursor: pointer; font-size: 13px; font-weight: 600; color: #666; border-bottom: 2px solid transparent; margin-bottom: -2px; transition: all .2s; }
.tab:hover { color: #2F5496; }
.tab.active { color: #2F5496; border-bottom-color: #2F5496; }
.panel { display: none; }
.panel.active { display: block; }
.kpi-row { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 12px; margin-bottom: 20px; }
.kpi { background: #fff; border-radius: 8px; padding: 14px; box-shadow: 0 1px 3px rgba(0,0,0,.08); text-align: center; }
.kpi .value { font-size: 28px; font-weight: 700; color: #2F5496; }
.kpi .value.warn { color: #d97706; }
.kpi .value.bad { color: #dc2626; }
.kpi .label { font-size: 11px; color: #888; margin-top: 2px; text-transform: uppercase; letter-spacing: .5px; }
.chart-row { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 20px; }
.chart-box { background: #fff; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,.08); }
.chart-box h3 { font-size: 13px; color: #444; margin-bottom: 10px; }
.chart-container { position: relative; height: 220px; }
.card { background: #fff; border-radius: 8px; padding: 16px; box-shadow: 0 1px 3px rgba(0,0,0,.08); margin-bottom: 16px; }
.card h3 { font-size: 13px; color: #444; margin-bottom: 10px; }
.tree, .tree ul { list-style: none; padding-left: 18px; }
.tree > li { padding-left: 0; }
.tree li { position: relative; padding: 2px 0; font-size: 13px; }
.tree li::before { content: ""; position: absolute; left: -10px; top: 0; bottom: 0; border-left: 1px solid #e0e0e0; }
.tree .id { display: inline-block; min-width: 34px; color: #999; font-variant-numeric: tabular-nums; font-size: 11px; }
.tree .folder { font-weight: 600; color: #2F5496; }
.tree .retired { color: #aaa; text-decoration: line-through; }
.pill { display: inline-block; padding: 1px 7px; border-radius: 4px; font-size: 10px; font-weight: 600; margin-left: 6px; }
.pill.delegated { background: #e0e7ff; color: #4338ca; }
.pill.item { background: #f3e8ff; color: #7e22ce; }
.pill.tool { background: #dcfce7; color: #16a34a; }
.empty { color: #999; font-size: 13px; padding: 8px 0; }
.filters { display: flex; gap: 10px; margin-bottom: 14px; flex-wrap: wrap; align-items: center; }
.filters select { padding: 6px 10px; border: 1px solid #d0d0d0; border-radius: 6px; font-size: 12px; background: #fff; }
.filters label { font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: .5px; }
.filters button { padding: 6px 12px; border: 1px solid #d0d0d0; border-radius: 6px; font-size: 12px; background: #fff; cursor: pointer; }
.filters button:hover { background: #f0f0f0; }
.filters .count { font-size: 12px; color: #666; margin-left: auto; }
.finding { padding: 7px 10px; border-radius: 6px; margin-bottom: 6px; font-size: 12.5px; }
.finding.error { background: #fee2e2; color: #991b1b; }
.finding.warn { background: #fef3c7; color: #92400e; }
.finding .where { font-family: ui-monospace, Consolas, monospace; font-size: 11.5px; }
.wgroup { margin-bottom: 12px; }
.wgroup > summary { cursor: pointer; font-size: 12.5px; color: #92400e; background: #fef3c7; padding: 7px 10px; border-radius: 6px; }
.wgroup > div { padding: 6px 0 0 14px; }
.wgroup li { font-size: 12px; color: #666; list-style: none; padding: 2px 0; font-family: ui-monospace, Consolas, monospace; }
.clean { background: #dcfce7; color: #166534; padding: 10px 12px; border-radius: 6px; font-size: 13px; }
.gridjs-wrapper { font-size: 12px; }
.gridjs-th { font-size: 11px !important; }
@media (max-width: 900px) { .chart-row { grid-template-columns: 1fr; } }
</style>
</head>
<body>

<div class="header">
  <div>
    <h1>Artifact Register &mdash; {{SCOPE}}</h1>
    <div class="subtitle">TSP.5 &nbsp;&middot;&nbsp; {{COUNT}} artifacts &nbsp;&middot;&nbsp; generated {{GENERATED}}</div>
  </div>
</div>

<div class="tabs">
  <div class="tab active" onclick="showTab('overview')">Overview</div>
  <div class="tab" onclick="showTab('artifacts')">Artifacts</div>
  <div class="tab" onclick="showTab('containment')">Containment</div>
  <div class="tab" onclick="showTab('review')">Review queue</div>
  <div class="tab" onclick="showTab('findings')">Findings {{FINDING_BADGE}}</div>
</div>

<div id="overview" class="panel active">
  <div class="kpi-row" id="kpis"></div>
  <div class="chart-row">
    <div class="chart-box"><h3>By type</h3><div class="chart-container"><canvas id="typeChart"></canvas></div></div>
    <div class="chart-box"><h3>By status</h3><div class="chart-container"><canvas id="statusChart"></canvas></div></div>
    <div class="chart-box"><h3>Where things are</h3><div class="chart-container"><canvas id="locChart"></canvas></div></div>
  </div>
  <div class="card"><h3>Delegated &mdash; the register stops here and another tool governs the contents</h3><div id="delegatedGrid"></div></div>
</div>

<div id="artifacts" class="panel">
  <div class="filters">
    <label for="fStatus">Status</label><select id="fStatus"></select>
    <label for="fType">Type</label><select id="fType"></select>
    <label for="fLocation">Location</label><select id="fLocation"></select>
    <label for="fArea">Area of focus</label><select id="fArea"></select>
    <label for="fManaged">Managed by</label><select id="fManaged"></select>
    <button onclick="resetFilters()">Clear</button>
    <span class="count" id="filterCount"></span>
  </div>
  <div id="allGrid"></div>
</div>

<div id="findings" class="panel">
  <div class="card">
    <h3>Errors</h3>
    <div id="errorList"></div>
  </div>
  <div class="card">
    <h3>Warnings</h3>
    <div id="warningList"></div>
  </div>
  <div class="card">
    <h3>What was checked</h3>
    <div id="scope" style="font-size:12.5px;color:#555;line-height:1.6"></div>
  </div>
</div>

<div id="containment" class="panel">
  <div class="card"><h3>Digital</h3><ul class="tree" id="digitalTree"></ul></div>
  <div class="card"><h3>Physical</h3><ul class="tree" id="physicalTree"></ul></div>
</div>

<div id="review" class="panel">
  <div class="card"><h3>Active artifacts, oldest review first</h3><div id="reviewGrid"></div></div>
  <div class="card"><h3>Retired &mdash; confirm these are actually gone, then delete the rows</h3><div id="retiredGrid"></div></div>
</div>

<script>
const DATA = {{DATA}};
const STALE_YEARS = {{STALE}};

function showTab(id) {
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  event.currentTarget.classList.add('active');
}

const val = (r, k) => (r[k] === null || r[k] === undefined) ? '' : String(r[k]).trim();
const active = DATA.filter(r => val(r, 'Status') === 'Active');
const retired = DATA.filter(r => val(r, 'Status') === 'Retired');
const delegated = DATA.filter(r => val(r, 'Managed By'));
const folders = DATA.filter(r => val(r, 'Type') === 'Folder');
const untyped = DATA.filter(r => !val(r, 'Type'));

function ageYears(r) {
  const d = val(r, 'Last Reviewed');
  if (!d) return null;
  const then = new Date(d);
  if (isNaN(then)) return null;
  return (Date.now() - then.getTime()) / (365.25 * 24 * 3600 * 1000);
}
const stale = active.filter(r => { const a = ageYears(r); return a === null || a > STALE_YEARS; });

const kpis = [
  ['Artifacts', DATA.length, ''],
  ['Active', active.length, ''],
  ['Retired', retired.length, retired.length ? 'warn' : ''],
  ['Delegated', delegated.length, ''],
  ['Containers', folders.length, ''],
  ['Untyped', untyped.length, untyped.length ? 'warn' : ''],
  ['Stale review', stale.length, stale.length ? 'bad' : '']
];
document.getElementById('kpis').innerHTML = kpis.map(
  ([label, value, cls]) => '<div class="kpi"><div class="value ' + cls + '">' + value +
    '</div><div class="label">' + label + '</div></div>').join('');

function tally(rows, key) {
  const counts = {};
  rows.forEach(r => { const v = val(r, key) || '(blank)'; counts[v] = (counts[v] || 0) + 1; });
  return counts;
}
const PALETTE = ['#2F5496', '#4472C4', '#8FAADC', '#16a34a', '#d97706', '#dc2626', '#7e22ce', '#0891b2'];

function doughnut(id, counts) {
  new Chart(document.getElementById(id), {
    type: 'doughnut',
    data: { labels: Object.keys(counts),
            datasets: [{ data: Object.values(counts), backgroundColor: PALETTE }] },
    options: { responsive: true, maintainAspectRatio: false,
               plugins: { legend: { position: 'bottom', labels: { boxWidth: 12, font: { size: 11 } } } } }
  });
}
doughnut('typeChart', tally(DATA, 'Type'));
doughnut('statusChart', tally(DATA, 'Status'));

const locCounts = tally(DATA, 'Location');
const locSorted = Object.entries(locCounts).sort((a, b) => b[1] - a[1]).slice(0, 10);
new Chart(document.getElementById('locChart'), {
  type: 'bar',
  data: { labels: locSorted.map(e => e[0]),
          datasets: [{ data: locSorted.map(e => e[1]), backgroundColor: '#4472C4' }] },
  options: { indexAxis: 'y', responsive: true, maintainAspectRatio: false,
             plugins: { legend: { display: false } },
             scales: { x: { ticks: { precision: 0 } } } }
});

function grid(el, rows, columns) {
  if (!rows.length) { document.getElementById(el).innerHTML = '<div class="empty">Nothing here.</div>'; return; }
  new gridjs.Grid({
    columns: columns,
    data: rows.map(r => columns.map(c => val(r, c))),
    search: true, sort: true, resizable: true,
    pagination: { limit: 25 }
  }).render(document.getElementById(el));
}

const MAIN_COLS = ['ID', 'Name', 'Type', 'Location', 'Parent Digital', 'Parent Physical',
                   'Managed By', 'Area of Focus', 'Status', 'Last Reviewed'];

const FILTERS = [['fStatus', 'Status'], ['fType', 'Type'], ['fLocation', 'Location'],
                 ['fArea', 'Area of Focus'], ['fManaged', 'Managed By']];

FILTERS.forEach(([el, key]) => {
  const seen = new Map();
  DATA.forEach(r => { const v = val(r, key); seen.set(v, (seen.get(v) || 0) + 1); });
  const opts = [...seen.entries()]
    .filter(([v]) => v !== '')
    .sort((a, b) => a[0].localeCompare(b[0], undefined, { numeric: true }));
  const blanks = seen.get('') || 0;
  const sel = document.getElementById(el);
  sel.innerHTML = '<option value="">All</option>' +
    opts.map(([v, n]) => '<option value="' + v + '">' + v + ' (' + n + ')</option>').join('') +
    (blanks ? '<option value="__BLANK__">(blank) (' + blanks + ')</option>' : '');
  sel.addEventListener('change', applyFilters);
});

let mainGrid = null;
function applyFilters() {
  const active = FILTERS
    .map(([el, key]) => [key, document.getElementById(el).value])
    .filter(([, v]) => v !== '');
  const rows = DATA.filter(r => active.every(([key, want]) =>
    want === '__BLANK__' ? val(r, key) === '' : val(r, key) === want));

  document.getElementById('filterCount').textContent =
    rows.length + ' of ' + DATA.length + ' artifacts';

  const data = rows.map(r => MAIN_COLS.map(c => val(r, c)));
  if (mainGrid) {
    mainGrid.updateConfig({ data: data }).forceRender();
  } else {
    mainGrid = new gridjs.Grid({
      columns: MAIN_COLS, data: data, search: true, sort: true,
      resizable: true, pagination: { limit: 25 }
    });
    mainGrid.render(document.getElementById('allGrid'));
  }
}
function resetFilters() {
  FILTERS.forEach(([el]) => { document.getElementById(el).value = ''; });
  applyFilters();
}
applyFilters();

grid('delegatedGrid', delegated, ['ID', 'Name', 'Managed By', 'Location', 'Status']);
grid('retiredGrid', retired, ['ID', 'Name', 'Type', 'Location', 'Last Reviewed', 'Comments']);

const byAge = active.slice().sort((a, b) => {
  const x = ageYears(a), y = ageYears(b);
  if (x === null) return -1;
  if (y === null) return 1;
  return y - x;
});
grid('reviewGrid', byAge, ['ID', 'Name', 'Type', 'Location', 'Last Reviewed', 'Status']);

function buildTree(elId, parentKey) {
  const byId = {};
  DATA.forEach(r => { byId[val(r, 'ID')] = r; });
  const children = {};
  const roots = [];
  DATA.forEach(r => {
    const p = val(r, parentKey);
    if (!p || p === 'N/A' || p === '-') return;
    if (p === 'Main' || !byId[p]) { roots.push(r); return; }
    (children[p] = children[p] || []).push(r);
  });
  const numeric = (a, b) => (parseInt(val(a, 'ID'), 10) || 0) - (parseInt(val(b, 'ID'), 10) || 0);

  function render(rows) {
    return '<ul>' + rows.sort(numeric).map(r => {
      const id = val(r, 'ID'), type = val(r, 'Type');
      let cls = type === 'Folder' ? 'folder' : '';
      if (val(r, 'Status') === 'Retired') cls += ' retired';
      let pills = '';
      if (val(r, 'Managed By')) pills += '<span class="pill delegated">' + val(r, 'Managed By') + '</span>';
      if (type === 'Tool') pills += '<span class="pill tool">tool</span>';
      if (type === 'Item') pills += '<span class="pill item">item</span>';
      return '<li><span class="id">' + id + '</span><span class="' + cls + '">' +
             val(r, 'Name') + '</span>' + pills +
             (children[id] ? render(children[id]) : '') + '</li>';
    }).join('') + '</ul>';
  }
  const el = document.getElementById(elId);
  el.outerHTML = roots.length
    ? '<div>' + render(roots) + '</div>'
    : '<div class="empty">No ' + parentKey.replace('Parent ', '').toLowerCase() + ' containment recorded.</div>';
}
buildTree('digitalTree', 'Parent Digital');
buildTree('physicalTree', 'Parent Physical');

// --- findings -----------------------------------------------------------
const FINDINGS = {{FINDINGS}};

const errEl = document.getElementById('errorList');
errEl.innerHTML = FINDINGS.errors.length
  ? FINDINGS.errors.map(m => '<div class="finding error"><span class="where">' +
      m.replace(/&/g, '&amp;').replace(/</g, '&lt;') + '</span></div>').join('')
  : '<div class="clean">No errors. The register is internally consistent' +
    (FINDINGS.checked.disk ? ', and matches the folder on disk.' : '.') + '</div>';

const warnEl = document.getElementById('warningList');
const groups = Object.entries(FINDINGS.warnings).sort((a, b) => b[1].length - a[1].length);
warnEl.innerHTML = groups.length
  ? groups.map(([kind, items]) =>
      '<details class="wgroup"><summary>' + kind + ' (' + items.length + ')</summary><div><ul>' +
      items.map(m => '<li>' + m.replace(/&/g, '&amp;').replace(/</g, '&lt;') + '</li>').join('') +
      '</ul></div></details>').join('')
  : '<div class="clean">No warnings.</div>';

document.getElementById('scope').innerHTML = [
  '<b>Register</b> &mdash; duplicate IDs, vocabulary, parents that point nowhere or at themselves, review age.',
  FINDINGS.checked.tsp
    ? '<b>Delegation</b> &mdash; every <code>Managed By</code> resolved against the TSP.3 tool register.'
    : '<b>Delegation</b> &mdash; <i>not checked.</i> Re-run with <code>--tsp-register</code>.',
  FINDINGS.checked.disk
    ? '<b>Disk</b> &mdash; reconciled against <code>' + FINDINGS.checked.disk +
      '</code>. The scan stops where the register stops: at delegated artifacts, at tools, ' +
      'and at containers the register names no children inside.'
    : '<b>Disk</b> &mdash; <i>not checked.</i> Re-run with <code>--root</code> to reconcile ' +
      'the register against the folder it describes.'
].map(s => '<div>' + s + '</div>').join('');
</script>
</body>
</html>
"""


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("register")
    p.add_argument("output")
    p.add_argument("--scope", help="name shown in the title (default: from the register)")
    p.add_argument("--root", help="folder this register describes; enables the disk "
                                  "reconciliation on the Findings tab")
    p.add_argument("--tsp-register", help="the TSP register; enables the "
                                          "delegation check on the Findings tab")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.register, data_only=True)
    ws = wb[S.SHEET] if S.SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    hrow, cols = S.find_header(ws)
    rows = S.read_rows(ws, hrow, cols)

    scope = args.scope
    if not scope:
        title = ws.cell(S.TITLE_ROW, 1).value or ws.cell(S.TITLE_ROW, 2).value
        scope = str(title).replace("Artifact Register - ", "").strip() if title else "Global"

    data = []
    for rec in rows:
        clean = {}
        for key, value in rec.items():
            if key == "_row":
                continue
            if isinstance(value, dt.datetime):
                clean[key] = value.date().isoformat()
            elif value is None:
                clean[key] = ""
            else:
                clean[key] = str(value).strip()
        data.append(clean)

    # Same checks the audit runs - one definition, so the dashboard and the
    # command line can never report different findings.
    errors, warnings, stats = checks.run(
        rows, root=args.root, tools=checks.load_tools(args.tsp_register))
    findings = {
        "errors": errors,
        "warnings": warnings,
        "checked": {"disk": (stats["disk"] or {}).get("root") if stats["disk"] else None,
                    "tsp": stats.get("delegations_resolved") is not None},
    }
    warning_count = sum(len(v) for v in warnings.values())
    badge = ""
    if errors:
        badge = '<span class="pill" style="background:#fee2e2;color:#dc2626">%d</span>' % len(errors)
    elif warning_count:
        badge = '<span class="pill" style="background:#fef3c7;color:#d97706">%d</span>' % warning_count

    html = (TEMPLATE
            .replace("{{SCOPE}}", scope)
            .replace("{{COUNT}}", str(len(data)))
            .replace("{{GENERATED}}", dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
            .replace("{{STALE}}", str(STALE_YEARS))
            .replace("{{FINDING_BADGE}}", badge)
            .replace("{{FINDINGS}}", json.dumps(findings, ensure_ascii=False))
            .replace("{{DATA}}", json.dumps(data, ensure_ascii=False)))

    with io.open(args.output, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(html)

    active = sum(1 for d in data if d.get("Status") == "Active")
    print("Dashboard generated: %s" % args.output)
    print("  Scope: %s | Artifacts: %d | Active: %d | Delegated: %d"
          % (scope, len(data), active, sum(1 for d in data if d.get("Managed By"))))
    print("  Findings: %d error(s), %d warning(s) in %d class(es)"
          % (len(errors), warning_count, len(warnings)))


if __name__ == "__main__":
    main()
