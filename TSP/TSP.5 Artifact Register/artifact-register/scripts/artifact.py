#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Row-level edits that must keep the register and the disk in step.

    python artifact.py add    <register> <path> --name "Prima Contract" --type Document
    python artifact.py add    <register> <path> --id 22        # existing row, lost its prefix
    python artifact.py rename <register> --id 22 --name "New Name" --root <folder>
    python artifact.py move   <register> --id 22 --parent-digital 12 --root <folder>
    python artifact.py retire <register> --id 22 --root <folder> --yes

`Name` and the parent columns are encoded into the filesystem, so changing either
in the spreadsheet alone leaves the register and the disk disagreeing in a way no
check can see — the audit matches on ID, not on name. These four operations change
both together.

Everything else - Last Reviewed, Description, Comments, Owner, Managed By, Area of
Focus, Type - is a plain field edit. Do those in Excel; a command that only writes
one cell is a worse spreadsheet.

`retire` disposes of the artifact and keeps the row, so the ID stays spent and an
old reference still resolves. It moves the artifact to --archive if you give one,
deletes it otherwise. It refuses to orphan registered children, and does nothing
without --yes.
"""
import argparse
import datetime as dt
import os
import shutil
import sys
import time

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checks                                                   # noqa: E402
import schema as S                                              # noqa: E402


# --- helpers ----------------------------------------------------------------

def open_register(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[S.SHEET] if S.SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    hrow, cols = S.find_header(ws)
    return wb, ws, hrow, cols, S.id_width(wb)


def rows_of(ws, hrow, cols):
    return S.read_rows(ws, hrow, cols)


def row_for(rows, artifact_id):
    for rec in rows:
        if S.clean(rec["ID"]) == str(artifact_id):
            return rec
    sys.exit("No row with ID %s in this register." % artifact_id)


def next_id(rows):
    used = [int(S.clean(r["ID"])) for r in rows if S.clean(r["ID"]).isdigit()]
    return max(used) + 1 if used else 0


def strip_prefix(name):
    match = S.ID_PREFIX.match(name)
    return name[match.end():].strip() if match else name


def target_name(artifact_id, name, original, width):
    """`<ID>. <Name>`, ID zero-padded, keeping the extension for files."""
    ext = os.path.splitext(original)[1] if os.path.isfile(original) else ""
    return "%s. %s%s" % (S.format_id(artifact_id, width), name, ext)


def locate(root, artifact_id):
    """Where the artifact sits on disk, by its ID prefix. None if absent."""
    if not root:
        return None
    prefixed, _ = S.scan_folder(root)
    hits = prefixed.get(str(artifact_id), [])
    if not hits:
        return None
    # The shallowest hit is the artifact; anything deeper is inside it.
    hits.sort(key=lambda p: p.count("/"))
    return os.path.join(root, hits[0].replace("/", os.sep))


def children_of(rows, artifact_id):
    out = []
    for rec in rows:
        digital, physical = S.parent_of(rec)
        if str(artifact_id) in (digital, physical):
            out.append(rec)
    return out


def rename_on_disk(path, new_name, dry):
    dest = os.path.join(os.path.dirname(path), new_name)
    if os.path.abspath(dest) == os.path.abspath(path):
        print("  disk    already named %r" % new_name)
        return path
    if os.path.exists(dest):
        sys.exit("Refusing to overwrite an existing %s" % dest)
    print("  disk    %s -> %s" % (os.path.basename(path), new_name))
    if not dry:
        os.rename(path, dest)
    return dest


def report(register, root, tsp_register):
    """Re-check after the edit, so a half-done operation says so now."""
    wb = openpyxl.load_workbook(register, data_only=True)
    ws = wb[S.SHEET] if S.SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    hrow, cols = S.find_header(ws)
    errors, warnings, _ = checks.run(S.read_rows(ws, hrow, cols), root=root,
                                     tools=checks.load_tools(tsp_register),
                                     width=S.id_width(wb))
    print("\n  check   %d error(s), %d warning(s)"
          % (len(errors), sum(len(v) for v in warnings.values())))
    for msg in errors[:5]:
        print("          ERROR %s" % msg)
    if len(errors) > 5:
        print("          ... and %d more (run audit_artifact_register.py)" % (len(errors) - 5))


# --- commands ---------------------------------------------------------------

def cmd_add(args):
    wb, ws, hrow, cols, width = open_register(args.register)
    rows = rows_of(ws, hrow, cols)

    if not os.path.exists(args.path):
        sys.exit("%s does not exist." % args.path)
    original = os.path.basename(args.path)

    if args.id is not None:
        # The row already exists; the file simply lost its prefix.
        rec = row_for(rows, args.id)
        artifact_id = S.clean(rec["ID"])
        name = args.name or S.clean(rec.get("Name")) or strip_prefix(original)
        if args.name:
            ws.cell(rec["_row"], cols["Name"], args.name)
        print("Re-prefixing existing artifact %s (%s)" % (artifact_id, name))
        row = rec["_row"]
    else:
        artifact_id = str(next_id(rows))
        name = args.name or os.path.splitext(strip_prefix(original))[0]
        row = hrow + 1 + len(rows)
        print("Adding artifact %s (%s)" % (artifact_id, name))
        values = {
            "ID": int(artifact_id), "Name": name, "Description": args.description,
            "Type": args.type, "Location": args.location,
            "Parent Digital": args.parent_digital or S.NO_PARENT,
            "Parent Physical": args.parent_physical or S.NO_PARENT,
            "Managed By": args.managed_by, "Area of Focus": args.area,
            "Owner": args.owner, "Created On": dt.date.today(),
            "Status": "Active", "Last Reviewed": dt.date.today(),
        }
        for key, value in values.items():
            if value is not None and key in cols:
                ws.cell(row, cols[key], value)
        print("  row     %d written" % row)

    rename_on_disk(args.path, target_name(artifact_id, name, args.path, width), args.dry_run)

    if not args.dry_run:
        wb.save(args.register)
        report(args.register, args.root, args.tsp_register)
    else:
        print("\n  dry run - nothing written. Re-run without --dry-run.")
    return 0


def cmd_rename(args):
    wb, ws, hrow, cols, width = open_register(args.register)
    rows = rows_of(ws, hrow, cols)
    rec = row_for(rows, args.id)
    old = S.clean(rec.get("Name"))
    print("Renaming artifact %s: %r -> %r" % (args.id, old, args.name))
    ws.cell(rec["_row"], cols["Name"], args.name)
    print("  row     %d updated" % rec["_row"])

    path = locate(args.root, args.id)
    if path:
        rename_on_disk(path, target_name(args.id, args.name, path, width), args.dry_run)
    else:
        print("  disk    nothing found with prefix %s%s" %
              (args.id, "" if args.root else " (pass --root to rename the file too)"))

    if not args.dry_run:
        wb.save(args.register)
        report(args.register, args.root, args.tsp_register)
    else:
        print("\n  dry run - nothing written.")
    return 0


def cmd_move(args):
    if args.parent_digital is None and args.parent_physical is None:
        sys.exit("Give --parent-digital and/or --parent-physical.")
    wb, ws, hrow, cols, width = open_register(args.register)
    rows = rows_of(ws, hrow, cols)
    rec = row_for(rows, args.id)
    ids = set(S.clean(r["ID"]) for r in rows)

    print("Moving artifact %s (%s)" % (args.id, S.clean(rec.get("Name"))))
    for flag, column in (("parent_digital", "Parent Digital"),
                         ("parent_physical", "Parent Physical")):
        value = getattr(args, flag)
        if value is None:
            continue
        if value not in ids and value not in (S.ROOT_PARENT, S.NO_PARENT):
            sys.exit("%s %r is not an ID in this register." % (column, value))
        if value == str(args.id):
            sys.exit("%s cannot point at itself." % column)
        print("  row     %s: %r -> %r"
              % (column, S.clean(rec.get(column)) or S.NO_PARENT, value))
        ws.cell(rec["_row"], cols[column], value)

    # Only the digital parent has a filesystem meaning.
    if args.root and args.parent_digital is not None:
        path = locate(args.root, args.id)
        if not path:
            print("  disk    nothing found with prefix %s" % args.id)
        elif args.parent_digital in (S.ROOT_PARENT, S.NO_PARENT):
            dest_dir = args.root
        else:
            dest_dir = locate(args.root, args.parent_digital)
            if not dest_dir or not os.path.isdir(dest_dir):
                sys.exit("Parent %s is not a folder on disk; move the file by hand."
                         % args.parent_digital)
        if path:
            dest = os.path.join(dest_dir, os.path.basename(path))
            if os.path.abspath(dest) == os.path.abspath(path):
                print("  disk    already in place")
            elif os.path.exists(dest):
                sys.exit("Refusing to overwrite %s" % dest)
            else:
                print("  disk    %s -> %s%s" % (os.path.basename(path),
                                                os.path.relpath(dest_dir, args.root),
                                                os.sep))
                if not args.dry_run:
                    shutil.move(path, dest)

    if not args.dry_run:
        wb.save(args.register)
        report(args.register, args.root, args.tsp_register)
    else:
        print("\n  dry run - nothing written.")
    return 0


def cmd_retire(args):
    wb, ws, hrow, cols, width = open_register(args.register)
    rows = rows_of(ws, hrow, cols)
    rec = row_for(rows, args.id)
    name = S.clean(rec.get("Name"))

    if S.clean(rec.get("Status")) == "Retired":
        print("Artifact %s (%s) is already retired." % (args.id, name))

    live = [r for r in children_of(rows, args.id)
            if S.clean(r.get("Status")) != "Retired"]
    if live:
        listing = ", ".join("%s %s" % (S.clean(r["ID"]), S.clean(r.get("Name")))
                            for r in live[:6])
        sys.exit("Refusing to retire %s (%s): %d active artifact(s) are inside it.\n"
                 "  %s%s\nRetire those first - deleting the container would take them "
                 "with it and leave rows pointing at nothing."
                 % (args.id, name, len(live), listing,
                    "" if len(live) <= 6 else ", ..."))

    print("Retiring artifact %s (%s)" % (args.id, name))
    verb = "archive" if args.archive else "delete"
    path = locate(args.root, args.id)
    if path:
        kind = "folder" if os.path.isdir(path) else "file"
        where = (" -> %s" % args.archive) if args.archive else ""
        print("  %-7s %s %s%s" % (verb, kind, os.path.relpath(path, args.root), where))
    elif args.root:
        print("  %-7s nothing on disk with prefix %s (physical artifact, or "
              "already gone)" % (verb, args.id))
    else:
        print("  %-7s skipped - pass --root to act on the file" % verb)

    print("  row     Status -> Retired (row %d kept; ID %s stays spent)"
          % (rec["_row"], args.id))

    if not args.yes:
        print("\n  Nothing done. This deletes the artifact - re-run with --yes.")
        return 1

    if path:
        if os.path.isdir(path):
            shutil.rmtree(path)
        else:
            os.remove(path)
    ws.cell(rec["_row"], cols["Status"], "Retired")
    ws.cell(rec["_row"], cols["Last Reviewed"], dt.date.today())
    if args.reason and "Comments" in cols:
        existing = S.clean(rec.get("Comments"))
        ws.cell(rec["_row"], cols["Comments"],
                ("%s | " % existing if existing else "") +
                "Retired %s: %s" % (dt.date.today().isoformat(), args.reason))
    wb.save(args.register)
    report(args.register, args.root, args.tsp_register)
    return 0


def cmd_repad(args):
    """Bring every filename up to the register's ID width.

    Needed once when a register adopts padding, and again if the width is raised
    because the artifact count outgrew it. Only the number changes: the name,
    separator and extension are left exactly as they are.
    """
    wb, ws, hrow, cols, width = open_register(args.register)
    rows = rows_of(ws, hrow, cols)

    if args.width:
        width = args.width
        if S.SHEET_SETTINGS in wb.sheetnames:
            st = wb[S.SHEET_SETTINGS]
            for r in range(1, st.max_row + 1):
                if str(st.cell(r, 1).value or "").strip().lower() == "id width":
                    st.cell(r, 2, width)
                    break
            else:
                st.cell(st.max_row + 1, 1, "ID width"); st.cell(st.max_row, 2, width)
        else:
            st = wb.create_sheet(S.SHEET_SETTINGS)
            st.cell(1, 1, "Setting"); st.cell(1, 2, "Value")
            st.cell(2, 1, "ID width"); st.cell(2, 2, width)
        print("Setting ID width to %d" % width)

    widest = max((len(S.clean(r["ID"])) for r in rows), default=1)
    if widest > width:
        sys.exit("This register has %d-digit IDs but the width is %d. "
                 "Re-run with --width %d, or padding would sort wrongly "
                 "above 10^%d." % (widest, width, widest, width))

    plan = []
    if args.root:
        prefixed, _ = S.scan_folder(args.root)
        known = set(S.clean(r["ID"]) for r in rows)
        for rid in sorted(prefixed, key=lambda k: int(k)):
            if rid not in known:
                continue
            for rel in prefixed[rid]:
                path = os.path.join(args.root, rel.replace("/", os.sep))
                current = os.path.basename(path)
                match = S.ID_PREFIX.match(current)
                sep = current[match.end(1):match.end()]
                want = "%s%s%s" % (S.format_id(rid, width), sep, current[match.end():])
                if current != want:
                    plan.append((os.path.abspath(path), want))

    for path, want in plan:
        print("  %s -> %s" % (os.path.basename(path), want))
    print("\n  %d to rename, %d already correct"
          % (len(plan), sum(len(v) for v in
                            (prefixed.values() if args.root else [])) - len(plan)))

    if args.dry_run:
        print("  dry run - nothing written.")
        return 0

    # Write and release the workbook before touching the filesystem: Windows will
    # not rename a directory that contains an open file, and the register usually
    # lives inside one of the folders being renamed.
    wb.save(args.register)
    wb.close()

    register = os.path.abspath(args.register)
    failed = []
    # Deepest first, so a parent rename never invalidates a path still queued.
    for path, want in sorted(plan, key=lambda pw: pw[0].count(os.sep), reverse=True):
        dest = os.path.join(os.path.dirname(path), want)
        for _ in range(5):
            try:
                os.rename(path, dest)
                if register == path or register.startswith(path + os.sep):
                    register = dest + register[len(path):]
                break
            except OSError:
                time.sleep(0.2)
        else:
            failed.append((os.path.basename(path), want))

    if failed:
        print("\n  Could not rename %d item(s) - close anything open "
              "inside them and re-run:" % len(failed))
        for old, want in failed:
            print("    %s  ->  %s" % (old, want))
    if register != os.path.abspath(args.register):
        print("  register is now %s" % os.path.relpath(register, args.root or "."))
    report(register, args.root, args.tsp_register)
    return 0


# --- entry point ------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    subs = p.add_subparsers(dest="command")

    def shared(sub, dry=True):
        sub.add_argument("register")
        sub.add_argument("--root", help="folder the register describes")
        sub.add_argument("--tsp-register", help="tool register, for the post-edit check")
        if dry:
            sub.add_argument("--dry-run", action="store_true",
                             help="show what would change, write nothing")

    s = subs.add_parser("add", help="register a new artifact and prefix its file")
    shared(s)
    s.add_argument("path", help="the file or folder to register")
    s.add_argument("--id", help="existing row whose file lost its prefix")
    s.add_argument("--name", help="default: the filename without its extension")
    s.add_argument("--type", choices=S.TYPES)
    s.add_argument("--description")
    s.add_argument("--location")
    s.add_argument("--parent-digital")
    s.add_argument("--parent-physical")
    s.add_argument("--managed-by", help="TSP.n of the tool governing its contents")
    s.add_argument("--area")
    s.add_argument("--owner")
    s.set_defaults(func=cmd_add)

    s = subs.add_parser("rename", help="change the name in the register and on disk")
    shared(s)
    s.add_argument("--id", required=True)
    s.add_argument("--name", required=True)
    s.set_defaults(func=cmd_rename)

    s = subs.add_parser("move", help="change the containing artifact, and move the file")
    shared(s)
    s.add_argument("--id", required=True)
    s.add_argument("--parent-digital")
    s.add_argument("--parent-physical")
    s.set_defaults(func=cmd_move)

    s = subs.add_parser("repad", help="pad every filename to the register's ID width")
    shared(s)
    s.add_argument("--width", type=int, help="also change the register's ID width")
    s.set_defaults(func=cmd_repad)

    s = subs.add_parser("retire", help="delete the artifact; keep the row and the ID")
    shared(s, dry=False)
    s.add_argument("--id", required=True)
    s.add_argument("--reason", help="recorded in Comments")
    s.add_argument("--archive", metavar="DIR",
                   help="move the artifact here instead of deleting it")
    s.add_argument("--yes", action="store_true",
                   help="required - without --archive this deletes files")
    s.set_defaults(func=cmd_retire)

    args = p.parse_args()
    if not getattr(args, "command", None):
        p.print_help()
        return 0
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
