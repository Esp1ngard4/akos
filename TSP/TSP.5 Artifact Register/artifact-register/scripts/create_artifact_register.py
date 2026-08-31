#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create an empty Artifact Register for a scope.

    python create_artifact_register.py "7. Artifact Register, Atlas.xlsx" "Atlas"

Builds the workbook programmatically rather than copying a template, so the
schema has exactly one definition and a fresh register can never carry another
scope's rows across.

The filename should start with the ID this register holds in its own scope -
it is an artifact of that scope like any other.
"""
import argparse
import datetime as dt
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import schema as S                                              # noqa: E402

HEAD_FILL = PatternFill("solid", fgColor="2F5496")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="2F5496")

WIDTHS = {"ID": 6, "Name": 34, "Description": 46, "Type": 12, "Location": 20,
          "Parent Digital": 14, "Parent Physical": 15, "Managed By": 12,
          "Area of Focus": 14, "Owner": 12, "Created On": 12, "Status": 10,
          "Last Reviewed": 14, "Comments": 40}

DEFAULT_LOCATIONS = [("Physical", "Home"), ("Physical", "Office"),
                     ("Digital", "Cloud drive"), ("Digital", "Local disk")]


def build(path, scope, locations, areas, force):
    if os.path.exists(path) and not force:
        raise SystemExit("%s already exists. Pass --force to overwrite." % path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = S.SHEET

    ws.cell(S.TITLE_ROW, 1, "Artifact Register - %s" % scope).font = TITLE_FONT
    ws.cell(S.TITLE_ROW + 1, 1,
            "Created %s. The ID is written onto the artifact itself."
            % dt.date.today().isoformat())

    for i, name in enumerate(S.COLUMNS, start=1):
        cell = ws.cell(S.HEADER_ROW, i, name)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = WIDTHS.get(name, 16)
    ws.freeze_panes = ws.cell(S.HEADER_ROW + 1, 1)

    cols = dict((n, i) for i, n in enumerate(S.COLUMNS, start=1))
    last = S.HEADER_ROW + 400

    def dropdown(column, values):
        letter = ws.cell(1, cols[column]).column_letter
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(values),
                            allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add("%s%d:%s%d" % (letter, S.HEADER_ROW + 1, letter, last))

    dropdown("Type", S.TYPES)
    dropdown("Status", S.STATUSES)
    if locations:
        dropdown("Location", [n for _, n in locations])

    loc = wb.create_sheet(S.LOOKUP_LOCATIONS)
    for i, name in enumerate(["ID", "Location", "Active", "Kind"], start=1):
        c = loc.cell(2, i, name); c.fill, c.font = HEAD_FILL, HEAD_FONT
    for i, (kind, name) in enumerate(locations, start=1):
        loc.cell(2 + i, 1, i); loc.cell(2 + i, 2, name)
        loc.cell(2 + i, 3, "Yes"); loc.cell(2 + i, 4, kind)
    loc.column_dimensions["B"].width = 24

    af = wb.create_sheet(S.LOOKUP_AREAS)
    for i, name in enumerate(["ID", "Name"], start=1):
        c = af.cell(2, i, name); c.fill, c.font = HEAD_FILL, HEAD_FONT
    for i, (aid, name) in enumerate(areas, start=1):
        af.cell(2 + i, 1, aid); af.cell(2 + i, 2, name)
    af.column_dimensions["B"].width = 24

    # The register is an artifact of its own scope, so it holds a row like
    # anything else. Its ID comes from the prefix on its own filename, which is
    # the convention this tool exists to enforce - stated by demonstration.
    match = S.ID_PREFIX.match(os.path.basename(path))
    seed = match.group(1) if match else "0"
    first = S.HEADER_ROW + 1
    ws.cell(first, cols["ID"], int(seed))
    ws.cell(first, cols["Name"], os.path.splitext(os.path.basename(path))[0])
    ws.cell(first, cols["Description"], "This register.")
    ws.cell(first, cols["Type"], "Document")
    ws.cell(first, cols["Parent Digital"], S.ROOT_PARENT)
    ws.cell(first, cols["Parent Physical"], S.NO_PARENT)
    ws.cell(first, cols["Status"], "Active")
    ws.cell(first, cols["Created On"], dt.date.today())
    ws.cell(first, cols["Last Reviewed"], dt.date.today())

    wb.save(path)
    return len(S.COLUMNS), seed


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("output", help="path of the register to create")
    p.add_argument("scope", help="what this register covers, e.g. a project or AF name")
    p.add_argument("--location", action="append", default=[],
                   metavar="Kind:Name", help="seed a location, e.g. Physical:Garage "
                                             "(repeatable; defaults are used if omitted)")
    p.add_argument("--area", action="append", default=[], metavar="ID:Name",
                   help="seed an Area of Focus, e.g. 5:Financial Freedom (repeatable)")
    p.add_argument("--force", action="store_true", help="overwrite an existing file")
    args = p.parse_args()

    locations = []
    for spec in args.location:
        kind, _, name = spec.partition(":")
        if kind.strip().capitalize() not in S.KINDS or not name:
            raise SystemExit("--location wants Physical:Name or Digital:Name, got %r" % spec)
        locations.append((kind.strip().capitalize(), name.strip()))
    areas = []
    for spec in args.area:
        aid, _, name = spec.partition(":")
        if not name:
            raise SystemExit("--area wants ID:Name, got %r" % spec)
        areas.append((aid.strip(), name.strip()))

    n, seed = build(args.output, args.scope, locations or DEFAULT_LOCATIONS,
                    areas, args.force)
    print("Created %s" % args.output)
    print("  scope: %s | %d columns | %d locations | %d areas"
          % (args.scope, n, len(locations or DEFAULT_LOCATIONS), len(areas)))
    print("  seeded with itself as artifact %s" % seed)
    print("\nNext: add artifacts, then")
    print("  python refresh_artifact_register.py \"%s\" \"Artifact Dashboard.html\"" % args.output)


if __name__ == "__main__":
    main()
