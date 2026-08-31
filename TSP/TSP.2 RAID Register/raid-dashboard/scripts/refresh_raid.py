#!/usr/bin/env python3
"""
Read a RAID .xlsx file and generate a self-contained HTML dashboard.
Usage: python refresh_raid.py <xlsx_path> <output_html_path> [project_name]
"""
import sys, json, os
from datetime import datetime, date

def read_raid_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Ticket Tracker"]

    # Find header row (row 6 by default, but scan for "RAID.ID")
    header_row = 6
    for r in range(1, 10):
        if ws.cell(row=r, column=2).value == "RAID.ID":
            header_row = r
            break

    # Find last data row
    last_row = header_row + 1
    while ws.cell(row=last_row, column=2).value is not None:
        last_row += 1
    last_row -= 1

    entries = []
    for row in range(header_row + 1, last_row + 1):
        def val(col):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (datetime, date)):
                return v.strftime("%Y-%m-%d")
            return v

        urgency = val(7) or 0
        consequences = val(8) or 0
        priority = round((urgency * 1.5 + consequences) / 12.5 * 100, 1) if (urgency or consequences) else 0

        probability = val(10) or 0
        severity = val(11) or 0
        mitigation_target = val(13)
        target_residual = round(probability * severity * (1 - mitigation_target / 100), 1) if (probability and severity and mitigation_target not in (None, "")) else None

        entry = {
            "id": val(2), "detail": val(3), "type": val(4), "dri": val(5),
            "priority": priority, "urgency": urgency, "consequences": consequences,
            "feasibility": val(9),
            "probability": val(10), "severity": val(11), "responseStrategy": val(12),
            "mitigationTarget": val(13), "targetResidualRisk": target_residual, "residualRisk": val(15),
            "moscow": val(16), "status": val(17),
            "lastReview": val(18), "reviewOn": val(19), "nextReviewOn": val(20),
            "description": val(21), "actionPlan": val(22),
            "acceptanceCriteria": val(23), "actionLog": val(24),
            "category": val(25), "trackedExternally": val(26), "openedOn": val(27), "requestedBy": val(28),
            "involve": val(29), "hasAuxMat": val(30),
            "estimatedEffort": val(31), "etc": val(32), "etcRenegotiated": val(33),
            "closedOn": val(34), "closedBy": val(35),
        }
        entries.append({k: v for k, v in entry.items() if v is not None})

    wb.close()
    return entries

def detect_types(entries):
    """Detect which RAID types are present to adapt visuals."""
    types = set(e.get("type", "") for e in entries)
    return {
        "has_risks": "Risk" in types,
        "has_actions": "Action" in types,
        "has_issues": "Issue" in types,
        "has_decisions": "Decision" in types,
        "has_ideas": "Idea" in types,
        "types_present": sorted(t for t in types if t),
    }

def generate_html(entries, project_name, type_info):
    data_json = json.dumps(entries, ensure_ascii=False)
    today = datetime.now().strftime("%Y-%m-%d")

    # Risk heatmap chart block (only if risks present)
    risk_chart_html = ""
    risk_chart_js = ""
    if type_info["has_risks"]:
        risk_chart_html = '<div class="chart-card"><h3>Risk Heat Map (Probability x Severity)</h3><div class="chart-container"><canvas id="chartRisk"></canvas></div></div>'
        risk_chart_js = """
    // Risk heat map: Probability x Severity, with green/amber/red zone background
    const risks = DATA.entries.filter(e => e.type === 'Risk' && e.status !== 'Resolved' && e.status !== 'Closed' && (e.probability||e.urgency) && (e.severity||e.consequences));
    if (risks.length >= 3 && document.getElementById('chartRisk')) {
      const heatmapBg = {
        id: 'heatmapBg',
        beforeDraw(chart) {
          const { ctx, chartArea, scales } = chart;
          if (!chartArea) return;
          for (let p = 1; p <= 5; p++) {
            for (let s = 1; s <= 5; s++) {
              const score = p * s;
              const color = score >= 15 ? 'rgba(220,38,38,0.15)' : score >= 6 ? 'rgba(245,158,11,0.15)' : 'rgba(34,197,94,0.15)';
              const x0 = scales.x.getPixelForValue(p - 0.5), x1 = scales.x.getPixelForValue(p + 0.5);
              const y0 = scales.y.getPixelForValue(s + 0.5), y1 = scales.y.getPixelForValue(s - 0.5);
              ctx.fillStyle = color;
              ctx.fillRect(x0, y0, x1 - x0, y1 - y0);
            }
          }
        }
      };
      new Chart(document.getElementById('chartRisk'), {
        type: 'scatter',
        data: { datasets: [{ label: 'Risks', data: risks.map(r => ({x: r.probability||r.urgency, y: r.severity||r.consequences, label: 'R.'+r.id+' — '+r.detail})), backgroundColor: '#dc2626', borderColor: '#991b1b', pointRadius: 7, pointHoverRadius: 9 }] },
        plugins: [heatmapBg],
        options: { responsive: true, maintainAspectRatio: false,
          scales: { x: {title:{display:true,text:'Probability'},min:0,max:6,ticks:{stepSize:1}}, y: {title:{display:true,text:'Severity'},min:0,max:6,ticks:{stepSize:1}} },
          plugins: { legend: {display:false}, tooltip: { callbacks: { label: ctx => ctx.raw.label + ' (P:'+ctx.raw.x+' S:'+ctx.raw.y+')' } } }
        }
      });
    }
"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RAID Dashboard — {project_name}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.5.0/dist/chart.umd.js" integrity="sha384-iU8HYtnGQ8Cy4zl7gbNMOhsDTTKX02BTXptVP/vqAWIaTfM7isw76iyZCsjL2eVi" crossorigin="anonymous"></script>
<style>
:root {{ color-scheme: light; }}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f8f9fa; color: #1a1a1a; padding: 16px; }}
.header {{ margin-bottom: 20px; }}
.header h1 {{ font-size: 20px; font-weight: 700; color: #111; }}
.header .meta {{ font-size: 12px; color: #666; margin-top: 4px; }}
.kpis {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-bottom: 20px; }}
.kpi {{ background: #fff; border-radius: 10px; padding: 14px; border: 1px solid #e5e7eb; }}
.kpi-value {{ font-size: 28px; font-weight: 700; }}
.kpi-label {{ font-size: 11px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 2px; }}
.kpi-red .kpi-value {{ color: #dc2626; }}
.kpi-amber .kpi-value {{ color: #d97706; }}
.kpi-green .kpi-value {{ color: #16a34a; }}
.kpi-blue .kpi-value {{ color: #2563eb; }}
.tabs {{ display: flex; gap: 4px; border-bottom: 2px solid #e5e7eb; margin-bottom: 16px; }}
.tab {{ padding: 8px 14px; cursor: pointer; font-size: 13px; font-weight: 500; color: #666; border-bottom: 2px solid transparent; margin-bottom: -2px; transition: all 0.15s; border-radius: 6px 6px 0 0; }}
.tab:hover {{ background: #f3f4f6; }}
.tab.active {{ color: #111; border-bottom-color: #2563eb; font-weight: 600; }}
.panel {{ display: none; }}
.panel.active {{ display: block; }}
.filters {{ display: flex; gap: 8px; margin-bottom: 14px; flex-wrap: wrap; }}
.filter-btn {{ padding: 5px 12px; border-radius: 16px; border: 1px solid #d1d5db; background: #fff; font-size: 12px; cursor: pointer; transition: all 0.15s; color: #444; }}
.filter-btn:hover {{ border-color: #9ca3af; }}
.filter-btn.active {{ background: #2563eb; color: #fff; border-color: #2563eb; }}
.table-wrap {{ overflow-x: auto; }}
table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
th {{ text-align: left; padding: 8px 10px; border-bottom: 2px solid #e5e7eb; color: #666; font-weight: 600; font-size: 11px; text-transform: uppercase; letter-spacing: 0.3px; cursor: pointer; white-space: nowrap; position: sticky; top: 0; background: #fff; }}
th:hover {{ color: #111; }}
td {{ padding: 8px 10px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }}
tr:hover td {{ background: #f8fafc; }}
.badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px; font-weight: 600; }}
.badge-issue {{ background: #fee2e2; color: #991b1b; }}
.badge-action {{ background: #dbeafe; color: #1e40af; }}
.badge-risk {{ background: #fef3c7; color: #92400e; }}
.badge-decision {{ background: #e0e7ff; color: #3730a3; }}
.badge-idea {{ background: #f0fdf4; color: #166534; }}
.status-open {{ color: #dc2626; font-weight: 600; }}
.status-resolved {{ color: #16a34a; font-weight: 600; }}
.status-closed {{ color: #16a34a; font-weight: 600; }}
.status-in\\ progress {{ color: #d97706; font-weight: 600; }}
.priority-bar {{ width: 60px; height: 6px; background: #f0f0f0; border-radius: 3px; overflow: hidden; display: inline-block; vertical-align: middle; margin-right: 6px; }}
.priority-fill {{ height: 100%; border-radius: 3px; }}
.detail-link {{ color: #2563eb; cursor: pointer; text-decoration: none; }}
.detail-link:hover {{ text-decoration: underline; }}
.modal-overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.3); z-index: 100; justify-content: center; align-items: flex-start; padding: 40px 16px; }}
.modal-overlay.open {{ display: flex; }}
.modal {{ background: #fff; border-radius: 12px; max-width: 560px; width: 100%; max-height: 80vh; overflow-y: auto; padding: 24px; box-shadow: 0 20px 60px rgba(0,0,0,0.15); }}
.modal h2 {{ font-size: 16px; margin-bottom: 12px; }}
.modal .field {{ margin-bottom: 10px; }}
.modal .field-label {{ font-size: 11px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 2px; }}
.modal .field-value {{ font-size: 13px; line-height: 1.5; white-space: pre-wrap; }}
.modal .close-btn {{ position: sticky; top: 0; float: right; background: #f3f4f6; border: none; border-radius: 6px; padding: 4px 10px; cursor: pointer; font-size: 13px; }}
.modal .close-btn:hover {{ background: #e5e7eb; }}
.modal .edit-btn {{ margin-top: 12px; padding: 8px 16px; background: #2563eb; color: #fff; border: none; border-radius: 8px; font-size: 13px; cursor: pointer; }}
.modal .edit-btn:hover {{ background: #1d4ed8; }}
.charts-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
.chart-card {{ background: #fff; border-radius: 10px; padding: 16px; border: 1px solid #e5e7eb; }}
.chart-card h3 {{ font-size: 13px; font-weight: 600; margin-bottom: 10px; color: #333; }}
.chart-container {{ position: relative; height: 200px; }}
.alert {{ padding: 10px 14px; border-radius: 8px; font-size: 12px; margin-bottom: 8px; display: flex; align-items: center; gap: 8px; }}
.alert-red {{ background: #fef2f2; border: 1px solid #fecaca; color: #991b1b; }}
.alert-amber {{ background: #fffbeb; border: 1px solid #fed7aa; color: #92400e; }}
</style>
</head>
<body>
<div class="header">
  <h1>RAID Dashboard — {project_name}</h1>
  <div class="meta">Last updated: {today}</div>
</div>
<div class="kpis" id="kpis"></div>
<div class="tabs" id="tabs">
  <div class="tab active" data-tab="board">Board</div>
  <div class="tab" data-tab="analytics">Analytics</div>
  <div class="tab" data-tab="reviews">Reviews</div>
</div>
<div class="panel active" id="panel-board">
  <div class="filters" id="filters"></div>
  <div class="table-wrap">
    <table>
      <thead><tr>
        <th data-sort="id">ID</th><th data-sort="detail">Detail</th><th data-sort="type">Type</th>
        <th data-sort="priority">Priority</th><th data-sort="moscow">MoSCoW</th>
        <th data-sort="status">Status</th><th data-sort="trackedExternally">Tracked</th><th data-sort="hasAuxMat">AuxMat</th><th data-sort="lastReview">Last Review</th>
      </tr></thead>
      <tbody id="tableBody"></tbody>
    </table>
  </div>
</div>
<div class="panel" id="panel-analytics">
  <div class="charts-grid">
    <div class="chart-card"><h3>By Type</h3><div class="chart-container"><canvas id="chartType"></canvas></div></div>
    <div class="chart-card"><h3>By Status</h3><div class="chart-container"><canvas id="chartStatus"></canvas></div></div>
    <div class="chart-card"><h3>Priority Distribution (Open)</h3><div class="chart-container"><canvas id="chartPriority"></canvas></div></div>
    <div class="chart-card"><h3>MoSCoW Breakdown (Open)</h3><div class="chart-container"><canvas id="chartMoscow"></canvas></div></div>
    {risk_chart_html}
    <div class="chart-card" style="grid-column: span 2"><h3>Items Over Time</h3><div class="chart-container"><canvas id="chartTimeline"></canvas></div></div>
  </div>
</div>
<div class="panel" id="panel-reviews"><div id="reviewAlerts"></div></div>
<div class="modal-overlay" id="modalOverlay"><div class="modal" id="modalContent"></div></div>

<script>
const DATA = {{entries: {data_json}}};
const FILTER_KEY = 'raidFilters_' + {json.dumps(project_name)};
let activeFilters = {{}}, sortCol = 'priority', sortDir = -1;
try {{ activeFilters = JSON.parse(localStorage.getItem(FILTER_KEY)) || {{}}; }} catch(e) {{ activeFilters = {{}}; }}
function clearFilters() {{
  activeFilters = {{}};
  try {{ localStorage.removeItem(FILTER_KEY); }} catch(e) {{}}
  renderFilters(); renderTable();
}}

function renderKPIs() {{
  const open = DATA.entries.filter(e => e.status && e.status !== 'Resolved' && e.status !== 'Closed');
  const critical = open.filter(e => e.priority >= 80);
  const resolved = DATA.entries.filter(e => e.status === 'Resolved' || e.status === 'Closed');
  const overdue = open.filter(e => e.reviewOn && new Date(e.reviewOn) < new Date());
  document.getElementById('kpis').innerHTML =
    '<div class="kpi kpi-red"><div class="kpi-value">'+open.length+'</div><div class="kpi-label">Open items</div></div>'+
    '<div class="kpi kpi-amber"><div class="kpi-value">'+critical.length+'</div><div class="kpi-label">Critical (>=80%)</div></div>'+
    '<div class="kpi kpi-green"><div class="kpi-value">'+resolved.length+'</div><div class="kpi-label">Resolved</div></div>'+
    '<div class="kpi kpi-blue"><div class="kpi-value">'+overdue.length+'</div><div class="kpi-label">Review overdue</div></div>';
}}

function renderFilters() {{
  const types = [...new Set(DATA.entries.map(e=>e.type).filter(Boolean))];
  const statuses = [...new Set(DATA.entries.map(e=>e.status).filter(Boolean))];
  const moscows = [...new Set(DATA.entries.map(e=>e.moscow).filter(Boolean))].sort();
  const categories = [...new Set(DATA.entries.map(e=>e.category).filter(Boolean))].sort();
  const dris = [...new Set(DATA.entries.map(e=>e.dri).filter(Boolean))].sort();
  let html = '<span style="font-size:11px;color:#666;padding:5px 0">Filter:</span>';
  types.forEach(t => html += '<button class="filter-btn" data-filter="type" data-val="'+t+'">'+t+'</button>');
  html += '<span style="color:#ddd">|</span>';
  statuses.forEach(s => html += '<button class="filter-btn" data-filter="status" data-val="'+s+'">'+s+'</button>');
  if (moscows.length > 1) {{
    html += '<span style="color:#ddd">|</span>';
    moscows.forEach(m => html += '<button class="filter-btn" data-filter="moscow" data-val="'+m+'">'+m.replace(/^\\d\\./,"")+'</button>');
  }}
  if (categories.length > 1) {{
    html += '<span style="color:#ddd">|</span>';
    categories.forEach(c => html += '<button class="filter-btn" data-filter="category" data-val="'+c+'">'+c+'</button>');
  }}
  if (dris.length > 1) {{
    html += '<span style="color:#ddd">|</span>';
    dris.forEach(d => html += '<button class="filter-btn" data-filter="dri" data-val="'+d+'">'+d+'</button>');
  }}
  html += '<span style="color:#ddd">|</span>';
  html += '<button class="filter-btn" data-filter="trackedExternally" data-val="Y">Tracked ✓</button>';
  html += '<button class="filter-btn" data-filter="trackedExternally" data-val="N">Tracked ✗</button>';
  const activeCount = Object.values(activeFilters).filter(Boolean).length;
  html += '<span style="color:#ddd">|</span>';
  html += '<button class="filter-btn" style="font-style:italic" onclick="clearFilters()">Clear all</button>';
  if (activeCount > 0) html += '<span style="font-size:11px;color:#2563eb;margin-left:4px">'+activeCount+' active</span>';
  document.getElementById('filters').innerHTML = html;
  document.querySelectorAll('.filter-btn[data-filter]').forEach(b => b.classList.toggle('active', activeFilters[b.dataset.filter] === b.dataset.val));
}}

document.getElementById('filters').addEventListener('click', e => {{
  const btn = e.target.closest('.filter-btn');
  if (!btn || !btn.dataset.filter) return;
  const key = btn.dataset.filter, val = btn.dataset.val;
  activeFilters[key] = activeFilters[key] === val ? null : val;
  try {{ localStorage.setItem(FILTER_KEY, JSON.stringify(activeFilters)); }} catch(e) {{}}
  renderFilters();
  renderTable();
}});

function renderTable() {{
  const items = DATA.entries.filter(e => {{
    for (const [k,v] of Object.entries(activeFilters)) {{ if (v && e[k] !== v) return false; }}
    return true;
  }}).sort((a, b) => {{
    let av = a[sortCol] ?? '', bv = b[sortCol] ?? '';
    if (typeof av === 'number' && typeof bv === 'number') return (av - bv) * sortDir;
    return String(av).localeCompare(String(bv)) * sortDir;
  }});
  document.getElementById('tableBody').innerHTML = items.map(e => {{
    const pColor = e.priority >= 80 ? '#dc2626' : e.priority >= 60 ? '#d97706' : '#16a34a';
    const typeClass = 'badge-'+(e.type||'').toLowerCase();
    const statusClass = 'status-'+(e.status||'').toLowerCase();
    return '<tr><td><strong>R.'+e.id+'</strong></td><td><a class="detail-link" onclick="showDetail('+e.id+')">'+e.detail+'</a></td><td><span class="badge '+typeClass+'">'+(e.type||'—')+'</span></td><td><div class="priority-bar"><div class="priority-fill" style="width:'+e.priority+'%;background:'+pColor+'"></div></div>'+e.priority+'%</td><td>'+(e.moscow||'—').replace(/^\\d\\./,'')+'</td><td><span class="'+statusClass+'">'+(e.status||'—')+'</span></td><td style="text-align:center">'+(e.trackedExternally==='Y'?'✓':'—')+'</td><td style="text-align:center">'+(e.hasAuxMat==='Y'?'📁':'—')+'</td><td>'+(e.lastReview||'—')+'</td></tr>';
  }}).join('');
}}

document.querySelector('thead').addEventListener('click', e => {{
  const th = e.target.closest('th'); if (!th) return;
  const col = th.dataset.sort;
  if (sortCol === col) sortDir *= -1; else {{ sortCol = col; sortDir = col === 'priority' ? -1 : 1; }}
  renderTable();
}});

function showDetail(id) {{
  const e = DATA.entries.find(x => x.id === id); if (!e) return;
  const riskFields = e.type === 'Risk' ? [['Probability',e.probability],['Severity',e.severity],['Response Strategy',e.responseStrategy],['Mitigation Target',e.mitigationTarget!=null?e.mitigationTarget+'%':null],['Target Residual Risk',e.targetResidualRisk],['Residual Risk Score',e.residualRisk]] : [];
  const fields = [['Type',e.type],['DRI',e.dri],['Priority',e.priority+'%'],['Urgency',e.urgency],['Consequences',e.consequences],['Feasibility',e.feasibility],...riskFields,['MoSCoW',e.moscow],['Status',e.status],['Tracked Externally',e.trackedExternally==='Y'?'Yes':'No'],['Description',e.description],['Action Plan',e.actionPlan],['Acceptance Criteria',e.acceptanceCriteria],['Action Log',e.actionLog],['Opened',e.openedOn],['Last Review',e.lastReview],['Review On',e.reviewOn],['Next Review On',e.nextReviewOn],['Closed On',e.closedOn],['Closed By',e.closedBy],['Requested By',e.requestedBy],['Estimated Effort',e.estimatedEffort],['ETC',e.etc],['ETC Renegotiated',e.etcRenegotiated],['Has AuxMat',e.hasAuxMat==='Y'?'Yes':'No']].filter(([,v])=>v!=null);
  document.getElementById('modalContent').innerHTML = '<button class="close-btn" onclick="closeModal()">x</button><h2>R.'+e.id+' — '+e.detail+'</h2>'+fields.map(([l,v])=>'<div class="field"><div class="field-label">'+l+'</div><div class="field-value">'+String(v).replace(/\\n/g,'<br>')+'</div></div>').join('')+'<button class="edit-btn" onclick="requestEdit('+e.id+')">Request Edit via Chat</button>';
  document.getElementById('modalOverlay').classList.add('open');
}}
function closeModal() {{ document.getElementById('modalOverlay').classList.remove('open'); }}
document.getElementById('modalOverlay').addEventListener('click', e => {{ if (e.target === e.currentTarget) closeModal(); }});
function requestEdit(id) {{
  const e = DATA.entries.find(x => x.id === id);
  const fn = window.cowork?.sendPrompt || window.sendPrompt;
  if (fn) fn('Update RAID item R.'+id+' ('+e.detail+'): [describe changes here]');
  closeModal();
}}

let chartsRendered = false;
function renderCharts() {{
  if (chartsRendered) return; chartsRendered = true;
  const open = DATA.entries.filter(e => e.status !== 'Resolved' && e.status !== 'Closed');
  const typeCounts = {{}};
  DATA.entries.forEach(e => {{ typeCounts[e.type||'Unknown'] = (typeCounts[e.type||'Unknown']||0) + 1; }});
  new Chart(document.getElementById('chartType'), {{ type:'doughnut', data:{{labels:Object.keys(typeCounts),datasets:[{{data:Object.values(typeCounts),backgroundColor:['#ef4444','#3b82f6','#f59e0b','#8b5cf6','#22c55e']}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}}}} }});
  const statusCounts = {{}};
  DATA.entries.forEach(e => {{ statusCounts[e.status||'Unknown'] = (statusCounts[e.status||'Unknown']||0) + 1; }});
  new Chart(document.getElementById('chartStatus'), {{ type:'doughnut', data:{{labels:Object.keys(statusCounts),datasets:[{{data:Object.values(statusCounts),backgroundColor:['#dc2626','#16a34a','#d97706','#9ca3af']}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}}}} }});
  const buckets = {{'Critical (>=80%)':0,'High (60-79%)':0,'Medium (40-59%)':0,'Low (<40%)':0}};
  open.forEach(e => {{ if(e.priority>=80)buckets['Critical (>=80%)']++;else if(e.priority>=60)buckets['High (60-79%)']++;else if(e.priority>=40)buckets['Medium (40-59%)']++;else buckets['Low (<40%)']++; }});
  new Chart(document.getElementById('chartPriority'), {{ type:'bar', data:{{labels:Object.keys(buckets),datasets:[{{data:Object.values(buckets),backgroundColor:['#dc2626','#f59e0b','#3b82f6','#22c55e']}}]}}, options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{stepSize:1}}}}}}}} }});
  const moscowCounts = {{}};
  open.forEach(e => {{ moscowCounts[e.moscow||'Unset'] = (moscowCounts[e.moscow||'Unset']||0) + 1; }});
  new Chart(document.getElementById('chartMoscow'), {{ type:'doughnut', data:{{labels:Object.keys(moscowCounts).map(m=>m.replace(/^\\d\\./,'')),datasets:[{{data:Object.values(moscowCounts),backgroundColor:['#dc2626','#f59e0b','#3b82f6','#9ca3af']}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}}}} }});
  {risk_chart_js}
  const events = [];
  DATA.entries.forEach(e => {{ if(e.openedOn)events.push({{date:e.openedOn,type:'opened'}}); if(e.closedOn)events.push({{date:e.closedOn,type:'closed'}}); }});
  events.sort((a,b)=>a.date.localeCompare(b.date));
  let cumOpen=0,cumClosed=0; const tlLabels=[],tlOpen=[],tlClosed=[];
  events.forEach(ev => {{ if(ev.type==='opened')cumOpen++;else cumClosed++; tlLabels.push(ev.date); tlOpen.push(cumOpen); tlClosed.push(cumClosed); }});
  new Chart(document.getElementById('chartTimeline'), {{ type:'line', data:{{labels:tlLabels,datasets:[{{label:'Opened',data:tlOpen,borderColor:'#dc2626',backgroundColor:'rgba(220,38,38,0.1)',fill:true,tension:0.3}},{{label:'Closed',data:tlClosed,borderColor:'#16a34a',backgroundColor:'rgba(22,163,74,0.1)',fill:true,tension:0.3}}]}}, options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}}}},scales:{{x:{{ticks:{{maxTicksLimit:8,font:{{size:10}}}}}}}}}} }});
}}

function renderReviews() {{
  const today = new Date(), open = DATA.entries.filter(e => e.status && e.status !== 'Resolved' && e.status !== 'Closed');
  const alerts = [];
  open.forEach(e => {{
    if (!e.lastReview) {{ alerts.push({{severity:'red',text:'R.'+e.id+' — '+e.detail+': Never reviewed'}}); return; }}
    const days = Math.floor((today - new Date(e.lastReview)) / 86400000);
    if (days > 90) alerts.push({{severity:'red',text:'R.'+e.id+' — '+e.detail+': Last reviewed '+days+' days ago ('+e.lastReview+')'}});
    else if (days > 30) alerts.push({{severity:'amber',text:'R.'+e.id+' — '+e.detail+': Last reviewed '+days+' days ago ('+e.lastReview+')'}});
  }});
  alerts.sort((a,b)=>(a.severity==='red'?0:1)-(b.severity==='red'?0:1));
  document.getElementById('reviewAlerts').innerHTML = alerts.length === 0
    ? '<p style="color:#16a34a;font-size:13px">All items reviewed within the last 30 days.</p>'
    : alerts.map(a=>'<div class="alert alert-'+a.severity+'"><span>'+(a.severity==='red'?'&#128308;':'&#128993;')+'</span><span>'+a.text+'</span></div>').join('');
}}

document.getElementById('tabs').addEventListener('click', e => {{
  const tab = e.target.closest('.tab'); if (!tab) return;
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  tab.classList.add('active');
  document.getElementById('panel-'+tab.dataset.tab).classList.add('active');
  if (tab.dataset.tab === 'analytics') renderCharts();
  if (tab.dataset.tab === 'reviews') renderReviews();
}});

renderKPIs(); renderFilters(); renderTable();
</script>
</body>
</html>"""
    return html

def main():
    if len(sys.argv) < 3:
        print("Usage: python refresh_raid.py <xlsx_path> <output_html_path> [project_name]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_path = sys.argv[2]
    project_name = sys.argv[3] if len(sys.argv) > 3 else os.path.basename(xlsx_path).replace("RAID ", "").replace(".xlsx", "")

    entries = read_raid_xlsx(xlsx_path)
    type_info = detect_types(entries)
    html = generate_html(entries, project_name, type_info)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"Dashboard generated: {output_path}")
    print(f"Entries: {len(entries)}, Types: {type_info['types_present']}")

if __name__ == "__main__":
    main()
